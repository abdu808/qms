from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
PLANS = ROOT / "ISO9001" / "الخطط والمشرات"
DATE = "2026-05-02"


ROWS = [
    {
        "domain": "الأثر الاجتماعي وملفات المستفيدين",
        "axis": "AXIS-01 — الأثر الاجتماعي والمستفيد",
        "strategic_goal": "STR-2026-003 — تعميق الأثر الاجتماعي",
        "operational_scope": "الوصول الخدمي، تحديث ملفات المستفيدين، قياس الأثر",
        "indicator_scope": "المستفيدون المخدومون، تغطية الاحتياج، مؤشرات الأثر",
        "initiative_scope": "مبادرات قياس الأثر وخدمة الأسر",
        "performance_owner": "إدارة الخدمة المجتمعية",
        "data_owner": "قسم البيانات والبحث الاجتماعي",
        "supporting_parties": "وحدة الاستراتيجية والتميز المؤسسي؛ إدارة المساعدات العينية عند وجود توزيع",
        "approval_party": "المدير التنفيذي / وحدة الاستراتيجية",
        "decision": "صحيح مع ضبط مصدر البيانات",
        "correction_note": "بيانات المستفيدين والاحتياج تصدر من الخدمة المجتمعية، ولا تعتمد من المستودع إلا في جانب التوزيع المنفذ.",
    },
    {
        "domain": "الكفالات والأيتام",
        "axis": "AXIS-01 — الأثر الاجتماعي والمستفيد",
        "strategic_goal": "STR-2026-003 — تعميق الأثر الاجتماعي",
        "operational_scope": "كفالة الأيتام واستمرارية الكافلين وتحسين أثر الكفالة",
        "indicator_scope": "عدد الأيتام المكفولين؛ استبقاء الكافلين؛ أثر/مصروف الكفالة",
        "initiative_scope": "INI-2026-005 — كفالة 850 يتيم مع زيادة سنوية ثابتة وتحسين أثر الكفالة",
        "performance_owner": "إدارة الخدمة المجتمعية",
        "data_owner": "قسم الدعم والرعاية",
        "supporting_parties": "تنمية الموارد للمانحين والكافلين؛ المالية للصرف والتسويات",
        "approval_party": "المدير التنفيذي",
        "decision": "يحتاج تعديل مالك",
        "correction_note": "المالك المؤسسي هو الخدمة المجتمعية، والشخص المسؤول الحالي خاتمة لا طلال. ينصح بفصل مؤشر استبقاء الكافلين عن مؤشر أثر/مصروف الكفالة لاحقًا.",
    },
    {
        "domain": "التمكين والتأهيل والتشغيل",
        "axis": "AXIS-01 — الأثر الاجتماعي والمستفيد",
        "strategic_goal": "STR-2026-004 — تمكين المستفيدين",
        "operational_scope": "حزم التمكين، التأهيل، تشغيل المستفيدين، المشاريع الصغيرة",
        "indicator_scope": "نسبة المستفيدين بحزمة تمكين؛ تشغيل المتدربين؛ استمرارية التشغيل",
        "initiative_scope": "INI-2026-008؛ INI-2026-009؛ مبادرات التدريب والتمكين",
        "performance_owner": "إدارة الخدمة المجتمعية",
        "data_owner": "قسم التمكين والتنمية",
        "supporting_parties": "مركز التدريب/المعهد عند التنفيذ التدريبي؛ تنمية الموارد عند وجود ممول",
        "approval_party": "المدير التنفيذي",
        "decision": "صحيح مع تنسيق تشغيلي",
        "correction_note": "تظل ملكية التمكين لدى الخدمة المجتمعية، ويكون مركز التدريب جهة تنفيذ أو شريك داخلي لا مالكًا للأثر الاجتماعي وحده.",
    },
    {
        "domain": "السلال والمساعدات العينية",
        "axis": "AXIS-01 — الأثر الاجتماعي والمستفيد",
        "strategic_goal": "STR-2026-006 — تحسين تجربة المستفيد",
        "operational_scope": "استلام وتخزين وتوزيع السلال والمساعدات العينية",
        "indicator_scope": "دقة الجرد؛ زمن تلبية طلبات الصرف؛ جودة التوزيع؛ الفاقد والتالف",
        "initiative_scope": "INI-2026-006 — توزيع السلال الغذائية",
        "performance_owner": "إدارة المساعدات العينية والمستودع",
        "data_owner": "قسم التوزيع والجرد",
        "supporting_parties": "الخدمة المجتمعية لقوائم الاستحقاق؛ المالية للقيد المحاسبي",
        "approval_party": "الخدمة المجتمعية + المدير التنفيذي",
        "decision": "صحيح مع فصل الاستحقاق عن التنفيذ",
        "correction_note": "طلال مناسب للتنفيذ العيني، لكن قرار الاستحقاق وقوائم المستفيدين تكون من الخدمة المجتمعية.",
    },
    {
        "domain": "رضا المستفيد وتجربة الخدمة",
        "axis": "AXIS-01 — الأثر الاجتماعي والمستفيد",
        "strategic_goal": "STR-2026-006 — تحسين تجربة المستفيد",
        "operational_scope": "الرضا، زمن الاستجابة، تجربة المستفيد، الشكاوى والملاحظات",
        "indicator_scope": "رضا المستفيد؛ الالتزام العام بـ SLA؛ معالجة الملاحظات",
        "initiative_scope": "INI-2026-032 — قياس تجربة المستفيد عبر QMS",
        "performance_owner": "إدارة الخدمة المجتمعية",
        "data_owner": "قسم البيانات والبحث الاجتماعي",
        "supporting_parties": "وحدة الاستراتيجية؛ تقنية المعلومات للنظام والاستبيانات",
        "approval_party": "وحدة الاستراتيجية",
        "decision": "صحيح مع ربط نظامي",
        "correction_note": "يجب أن يرتبط المؤشر باستبيان دوري داخل النظام أو قناة موثقة، مع دورة اعتماد واضحة.",
    },
    {
        "domain": "الاستثمار والمحفظة والعوائد",
        "axis": "AXIS-02 — الاستدامة المالية",
        "strategic_goal": "STR-2026-007 — الاستدامة المالية",
        "operational_scope": "المحفظة الاستثمارية، العقارات، الأصول، العوائد الاستثمارية",
        "indicator_scope": "ROI؛ نمو الأصول؛ نسبة الإشغال والتحصيل؛ مساهمة الاستثمار في الإيرادات",
        "initiative_scope": "INI-2026-003 — تخصيص بند سنوي لبناء محفظة استثمارية آمنة",
        "performance_owner": "وحدة الاستثمار",
        "data_owner": "وحدة الاستثمار + الإدارة المالية",
        "supporting_parties": "الإدارة المالية؛ مجلس الإدارة؛ المدير التنفيذي",
        "approval_party": "مجلس الإدارة / المدير التنفيذي",
        "decision": "يحتاج تثبيت ملكية",
        "correction_note": "لا توضع الاستثمارات عند تنمية الموارد إلا إذا كانت نادية مسؤولة رسميًا عن وحدة الاستثمار. التخصيص يكون من الإيرادات غير المقيدة وعوائد الاستثمار بعد مراجعة السيولة.",
    },
    {
        "domain": "التبرعات والحملات والمانحون",
        "axis": "AXIS-02 — الاستدامة المالية",
        "strategic_goal": "STR-2026-007 — الاستدامة المالية",
        "operational_scope": "الحملات، المانحون، التبرعات، تنمية الإيرادات غير الاستثمارية",
        "indicator_scope": "إجمالي التبرعات؛ تنوع مصادر الدخل؛ تحقيق مستهدفات الحملات",
        "initiative_scope": "INI-2026-004؛ حملات التبرع الموسمية",
        "performance_owner": "إدارة تنمية الموارد والمشاريع",
        "data_owner": "قسم تنمية الموارد والعلاقات مع المانحين",
        "supporting_parties": "الاتصال المؤسسي للحملات؛ المالية للمطابقة والقيود",
        "approval_party": "المدير التنفيذي",
        "decision": "صحيح",
        "correction_note": "نادية مناسبة هنا إذا كانت مسؤولة تنمية الموارد، مع عدم نقل ملكية الاستثمار لهذا المسار.",
    },
    {
        "domain": "استبقاء المانحين الكبار",
        "axis": "AXIS-02 — الاستدامة المالية",
        "strategic_goal": "STR-2026-007 — الاستدامة المالية",
        "operational_scope": "علاقات المانحين، المتابعة، تقارير الأثر، الاستبقاء السنوي",
        "indicator_scope": "نسبة استبقاء المانحين الكبار؛ قيمة المساهمات المتكررة",
        "initiative_scope": "INI-2026-011 — استبقاء 80% من المانحين الكبار",
        "performance_owner": "إدارة تنمية الموارد والمشاريع",
        "data_owner": "قسم العلاقات مع المانحين",
        "supporting_parties": "الاتصال المؤسسي؛ المالية؛ الخدمة المجتمعية لتقارير الأثر",
        "approval_party": "المدير التنفيذي",
        "decision": "صحيح",
        "correction_note": "الملكية علاقاتية وتنموية، وليست مالية بحتة؛ المالية تثبت الأرقام ولا تملك العلاقة.",
    },
    {
        "domain": "الشراكات والفعاليات",
        "axis": "AXIS-04 — الشراكات ورأس المال البشري",
        "strategic_goal": "STR-2026-016 — منظومة الشراكات",
        "operational_scope": "تطوير الشراكات، تفعيلها، الفعاليات، المتابعة",
        "indicator_scope": "عدد الشراكات الفعالة؛ قيمة الشراكات؛ نسبة تفعيل الاتفاقيات",
        "initiative_scope": "INI-2026-023 — مراجعة وتطوير الشراكات القائمة",
        "performance_owner": "إدارة الاتصال المؤسسي والشراكات",
        "data_owner": "قسم الشراكات والفعاليات",
        "supporting_parties": "تنمية الموارد عند وجود تمويل؛ الخدمة المجتمعية عند وجود أثر مستفيدين",
        "approval_party": "المدير التنفيذي",
        "decision": "صحيح",
        "correction_note": "فاطمة مناسبة للشراكات، مع عدم تحميل الإدارة مؤشرات مالية مباشرة إلا بقدر قيمة الشراكة المثبتة.",
    },
    {
        "domain": "الإعلام والنشر المؤسسي",
        "axis": "AXIS-04 — الشراكات ورأس المال البشري",
        "strategic_goal": "STR-2026-016 — منظومة الشراكات",
        "operational_scope": "الإعلام، العلاقات العامة، النشر، التقارير الإعلامية",
        "indicator_scope": "المواد الإعلامية؛ التغطيات؛ التفاعل الرقمي؛ النشر الرسمي",
        "initiative_scope": "INI-2026-031 — 24 مادة إعلامية موثقة",
        "performance_owner": "إدارة الاتصال المؤسسي والشراكات",
        "data_owner": "قسم الإعلام والعلاقات العامة",
        "supporting_parties": "وحدة الاستراتيجية للمحتوى المعتمد؛ الموقع الرسمي للجمعية للنشر",
        "approval_party": "المدير التنفيذي",
        "decision": "صحيح مع تعديل قناة النشر",
        "correction_note": "النشر يكون على الموقع الرسمي للجمعية وليس على البوابة العامة الملغاة.",
    },
    {
        "domain": "التطوع والمشاركة المجتمعية",
        "axis": "AXIS-04 — الشراكات ورأس المال البشري",
        "strategic_goal": "STR-2026-017 — الكوادر والتطوع",
        "operational_scope": "تصميم مسارات التطوع، جذب المتطوعين، تفعيل المشاركة المجتمعية",
        "indicator_scope": "ساعات التطوع؛ عدد المتطوعين النشطين؛ جودة المشاركة",
        "initiative_scope": "مبادرات التطوع والمشاركة ضمن الاتصال المؤسسي",
        "performance_owner": "إدارة الاتصال المؤسسي والشراكات",
        "data_owner": "قسم التطوع والمشاركة المجتمعية",
        "supporting_parties": "الموارد البشرية عند التدريب الداخلي؛ الخدمة المجتمعية عند خدمة المستفيدين",
        "approval_party": "المدير التنفيذي",
        "decision": "صحيح حسب الهيكل الحالي",
        "correction_note": "يبقى التطوع في الاتصال المؤسسي حسب الهيكل، مع تنسيق الموارد البشرية في التدريب والتهيئة.",
    },
    {
        "domain": "رأس المال البشري وتدريب الموظفين",
        "axis": "AXIS-04 — الشراكات ورأس المال البشري",
        "strategic_goal": "STR-2026-017 — الكوادر والتطوع",
        "operational_scope": "تدريب الموظفين، الشهادات، الرضا الوظيفي، مراجعات الأداء",
        "indicator_scope": "20 ساعة تدريب/موظف؛ نسبة خطط التدريب؛ رضا الموظفين؛ مراجعات الأداء",
        "initiative_scope": "INI-2026-027؛ INI-2026-028",
        "performance_owner": "إدارة الدعم المؤسسي",
        "data_owner": "قسم الموارد البشرية",
        "supporting_parties": "وحدة الاستراتيجية للتقارير؛ المدير التنفيذي للاعتماد",
        "approval_party": "المدير التنفيذي",
        "decision": "صحيح",
        "correction_note": "خليل مناسب إذا كان مسؤول الموارد البشرية أو التدريب الداخلي، ويجب فصل تدريب الموظفين عن تدريب المستفيدين.",
    },
    {
        "domain": "التقنية والتحول الرقمي",
        "axis": "AXIS-03 — التميز المؤسسي والتحول الرقمي",
        "strategic_goal": "STR-2026-013 — التحول الرقمي",
        "operational_scope": "تشغيل الأنظمة، التكاملات، الأتمتة، أمن المعلومات، الدعم الفني",
        "indicator_scope": "توافر الأنظمة؛ الأتمتة الرقمية؛ زمن الاستجابة؛ التكاملات",
        "initiative_scope": "INI-2026-017 — تعزيز النضج الرقمي وتكامل الأنظمة",
        "performance_owner": "إدارة الدعم المؤسسي",
        "data_owner": "قسم تقنية المعلومات",
        "supporting_parties": "وحدة الاستراتيجية؛ الإدارات المالكة للعمليات",
        "approval_party": "المدير التنفيذي + وحدة الاستراتيجية",
        "decision": "يحتاج مالك تشغيلي واضح",
        "correction_note": "لا يبقى المدير التنفيذي مالكًا تنفيذيًا دائمًا؛ التقنية تملك التنفيذ، ووحدة الاستراتيجية تملك المتابعة والمنهجية.",
    },
    {
        "domain": "الذكاء الاصطناعي والأتمتة",
        "axis": "AXIS-03 — التميز المؤسسي والتحول الرقمي",
        "strategic_goal": "STR-2026-013 — التحول الرقمي",
        "operational_scope": "أدوات AI، أتمتة العمليات، توفير ساعات العمل، رفع الكفاءة",
        "indicator_scope": "ساعات العمل الموفرة بـ AI؛ عدد العمليات المؤتمتة؛ تبني الأدوات",
        "initiative_scope": "INI-2026-018 — تدريب 30 موظفًا على 3 أدوات AI",
        "performance_owner": "تقنية المعلومات / الدعم المؤسسي",
        "data_owner": "قسم تقنية المعلومات",
        "supporting_parties": "الموارد البشرية للتدريب؛ وحدة الاستراتيجية للقياس",
        "approval_party": "وحدة الاستراتيجية + المدير التنفيذي",
        "decision": "يحتاج تثبيت مؤشر ومالك",
        "correction_note": "المؤشر استراتيجي، لكن تنفيذه تشغيلي داخل التقنية مع دعم الموارد البشرية للتدريب.",
    },
    {
        "domain": "ISO 9001 والجودة والتوثيق",
        "axis": "AXIS-03 — التميز المؤسسي والتحول الرقمي",
        "strategic_goal": "STR-2026-012 — التميز والجودة",
        "operational_scope": "ISO 9001، السياسات، الإجراءات، المراجعات، التحسين المستمر",
        "indicator_scope": "شهادات التميز؛ توثيق السياسات؛ إغلاق الملاحظات؛ مراجعات الجودة",
        "initiative_scope": "INI-2026-033 — مسار ISO 9001",
        "performance_owner": "وحدة الاستراتيجية والتميز المؤسسي",
        "data_owner": "قسم الجودة والتميز المؤسسي",
        "supporting_parties": "جميع الإدارات حسب الإجراء؛ الرقابة الداخلية عند الملاحظات",
        "approval_party": "المدير التنفيذي",
        "decision": "صحيح",
        "correction_note": "إيلاف مناسبة إذا كانت مسؤولة الجودة/التميز. الوحدة تملك الحوكمة والتوثيق ولا تتحول إلى منفذ لكل الإدارات.",
    },
    {
        "domain": "متابعة الخطة والمؤشرات",
        "axis": "جميع المحاور",
        "strategic_goal": "كل الأهداف الاستراتيجية",
        "operational_scope": "PMO، متابعة المؤشرات، التقارير، المراجعات الربع سنوية",
        "indicator_scope": "نسبة إنجاز المبادرات؛ اكتمال تحديث المؤشرات؛ جودة التقارير",
        "initiative_scope": "حوكمة الخطة ومتابعة التنفيذ",
        "performance_owner": "وحدة الاستراتيجية والتميز المؤسسي",
        "data_owner": "جميع الإدارات حسب المؤشر",
        "supporting_parties": "كل الإدارات؛ المدير التنفيذي",
        "approval_party": "المدير التنفيذي",
        "decision": "صحيح مع توضيح الدور",
        "correction_note": "الوحدة لا تنفذ بدل الإدارات، بل تراجع وتضبط التعاريف وتصدر التقارير وتدير دورات المراجعة.",
    },
    {
        "domain": "الموازنة والصرف والتقارير المالية",
        "axis": "AXIS-02 — الاستدامة المالية",
        "strategic_goal": "STR-2026-007 — الاستدامة المالية",
        "operational_scope": "الموازنة، الصرف، القيود، الإقفال، التقارير المالية، التدفقات",
        "indicator_scope": "كفاءة الصرف؛ الإقفال الشهري؛ دقة التقارير؛ المطابقات البنكية",
        "initiative_scope": "ترشيد المصروفات والضبط المالي",
        "performance_owner": "الإدارة المالية",
        "data_owner": "قسم المحاسبة والتقارير المالية",
        "supporting_parties": "المدير التنفيذي؛ وحدة الاستثمار؛ تنمية الموارد",
        "approval_party": "المدير التنفيذي / المجلس",
        "decision": "صحيح",
        "correction_note": "المالية تعتمد الأثر المالي وتثبت القيود، لكنها لا تملك كل مؤشرات الإيراد أو علاقات المانحين.",
    },
    {
        "domain": "الرقابة الداخلية والمخاطر",
        "axis": "AXIS-03 — التميز المؤسسي والتحول الرقمي",
        "strategic_goal": "STR-2026-012 — التميز والجودة",
        "operational_scope": "المراجعة الداخلية، الامتثال، المخاطر، إغلاق الملاحظات",
        "indicator_scope": "تنفيذ خطة المراجعة؛ إغلاق الملاحظات؛ تكرار الملاحظات؛ الامتثال",
        "initiative_scope": "حوكمة الالتزام والمخاطر",
        "performance_owner": "وحدة الرقابة الداخلية",
        "data_owner": "وحدة الرقابة الداخلية",
        "supporting_parties": "لجنة المراجعة؛ المدير التنفيذي؛ الإدارات المعنية بالملاحظات",
        "approval_party": "لجنة المراجعة",
        "decision": "صحيح",
        "correction_note": "تبقى الرقابة منفصلة عن وحدة الاستراتيجية، مع تكامل في التقارير دون خلط في الاستقلالية.",
    },
]


CORRECTION_DECISIONS = [
    "اعتماد وحدة الاستثمار مالكًا للاستثمارات والمحفظة والعوائد، وليس تنمية الموارد.",
    "اعتماد تنمية الموارد مالكًا للحملات والمانحين والتبرعات، وليس الاستثمار.",
    "اعتماد الخدمة المجتمعية مالكًا للكفالة، الأثر، الرضا، التمكين، وملفات المستفيدين.",
    "اعتماد المساعدات العينية والمستودع مالكًا للتوزيع والجرد والتنفيذ العيني فقط.",
    "اعتماد وحدة الاستراتيجية والتميز مالكًا لحوكمة المؤشرات والتقارير، لا مالكًا مباشرًا لكل النتائج.",
    "اعتماد تقنية المعلومات/الدعم المؤسسي مالكًا تنفيذيًا للتحول الرقمي والذكاء الاصطناعي.",
    "تقليل الملكيات المباشرة المسندة للمدير التنفيذي، وتحويلها إلى اعتماد ومتابعة عليا.",
]


ACCEPTANCE_TESTS = [
    "كل مؤشر له: مالك أداء، مالك بيانات، جهة اعتماد.",
    "لا يوجد مؤشر مالي أو استثماري مملوك لقسم غير مختص.",
    "لا يوجد مؤشر مستفيدين مملوك للمستودع.",
    "لا يوجد مؤشر جودة أو ISO بلا متابعة من وحدة الاستراتيجية والتميز.",
    "لا يبقى المدير التنفيذي مالكًا تشغيليًا إلا مؤقتًا أو للملفات العابرة للإدارات.",
    "يمكن إدخال هذه المصفوفة لاحقًا في النظام أو ملف Excel بدون إعادة تفسير.",
]


ASSUMPTIONS = [
    "الهيكل التنظيمي الحالي هو المرجع الرسمي.",
    "أسماء الأشخاص تربط لاحقًا حسب شاغلي المناصب، أما هذه المصفوفة فتعتمد الملكية المؤسسية أولًا.",
    "عند وجود تعارض بين الشخص والقسم، نعتمد القسم الصحيح ثم نربط الشخص المسؤول عنه.",
]


HEADERS = [
    ("domain", "المجال"),
    ("axis", "المحور الاستراتيجي"),
    ("strategic_goal", "الهدف الاستراتيجي"),
    ("operational_scope", "نطاق الهدف التشغيلي"),
    ("indicator_scope", "نطاق المؤشرات"),
    ("initiative_scope", "المبادرات المرتبطة"),
    ("performance_owner", "مالك الأداء"),
    ("data_owner", "مالك البيانات"),
    ("supporting_parties", "الجهات المساندة"),
    ("approval_party", "جهة الاعتماد"),
    ("decision", "الحكم"),
    ("correction_note", "ملاحظات التصحيح"),
]


def write_markdown(path: Path) -> None:
    lines = [
        "# مصفوفة ملكية الخطة والمؤشرات والأقسام",
        "",
        f"**التاريخ:** {DATE}",
        "",
        "## الملخص",
        "",
        "هذه المصفوفة تعتمد الهيكل التنظيمي الحالي مرجعًا، وتصحح الملكيات بين الخطة والمؤشرات والمبادرات حتى يكون لكل مؤشر مالك أداء ومالك بيانات وجهة اعتماد.",
        "",
        "## المصفوفة",
        "",
        "| المجال | المحور | مالك الأداء | مالك البيانات | جهة الاعتماد | الحكم | ملاحظات التصحيح |",
        "|---|---|---|---|---|---|---|",
    ]
    for row in ROWS:
        lines.append(
            "| "
            + " | ".join(
                [
                    row["domain"],
                    row["axis"],
                    row["performance_owner"],
                    row["data_owner"],
                    row["approval_party"],
                    row["decision"],
                    row["correction_note"],
                ]
            )
            + " |"
        )

    lines.extend(["", "## أهم قرارات التصحيح", ""])
    lines.extend(f"- {item}" for item in CORRECTION_DECISIONS)
    lines.extend(["", "## اختبار القبول", ""])
    lines.extend(f"- {item}" for item in ACCEPTANCE_TESTS)
    lines.extend(["", "## الافتراضات", ""])
    lines.extend(f"- {item}" for item in ASSUMPTIONS)
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def write_json(path: Path) -> None:
    payload = {
        "title": "مصفوفة ملكية الخطة والمؤشرات والأقسام",
        "date": DATE,
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "source": {
            "organizationStructure": r"C:\Users\abdu8\OneDrive\2025\لهيكل التنظيمي\الهيكل التنظيمي.pptx",
            "strategicPlanFolder": str(PLANS),
        },
        "rows": ROWS,
        "correctionDecisions": CORRECTION_DECISIONS,
        "acceptanceTests": ACCEPTANCE_TESTS,
        "assumptions": ASSUMPTIONS,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def apply_sheet_style(ws, freeze: str = "A2") -> None:
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = freeze
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
            cell.border = border

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font


def autosize(ws, widths: dict[int, int] | None = None) -> None:
    widths = widths or {}
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        if col_idx in widths:
            ws.column_dimensions[letter].width = widths[col_idx]
            continue
        max_len = 12
        for row_idx in range(1, min(ws.max_row, 100) + 1):
            value = ws.cell(row_idx, col_idx).value
            if value:
                max_len = max(max_len, min(len(str(value)), 55))
        ws.column_dimensions[letter].width = max_len + 3


def write_excel(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "مصفوفة الملكية"
    ws.append([label for _, label in HEADERS])
    for row in ROWS:
        ws.append([row[key] for key, _ in HEADERS])
    apply_sheet_style(ws)
    autosize(
        ws,
        {
            1: 28,
            2: 30,
            3: 30,
            4: 36,
            5: 34,
            6: 34,
            7: 28,
            8: 28,
            9: 34,
            10: 28,
            11: 22,
            12: 55,
        },
    )

    ws2 = wb.create_sheet("قرارات التصحيح")
    ws2.sheet_view.rightToLeft = True
    ws2.append(["م", "قرار التصحيح"])
    for i, item in enumerate(CORRECTION_DECISIONS, 1):
        ws2.append([i, item])
    apply_sheet_style(ws2)
    autosize(ws2, {1: 8, 2: 90})

    ws3 = wb.create_sheet("اختبار القبول")
    ws3.sheet_view.rightToLeft = True
    ws3.append(["م", "معيار القبول"])
    for i, item in enumerate(ACCEPTANCE_TESTS, 1):
        ws3.append([i, item])
    apply_sheet_style(ws3)
    autosize(ws3, {1: 8, 2: 90})

    ws4 = wb.create_sheet("الافتراضات")
    ws4.sheet_view.rightToLeft = True
    ws4.append(["م", "الافتراض"])
    for i, item in enumerate(ASSUMPTIONS, 1):
        ws4.append([i, item])
    apply_sheet_style(ws4)
    autosize(ws4, {1: 8, 2: 90})

    wb.save(path)


def main() -> None:
    DOCS.mkdir(parents=True, exist_ok=True)
    PLANS.mkdir(parents=True, exist_ok=True)

    md_path = DOCS / f"strategic-ownership-matrix-{DATE}.md"
    json_path = DOCS / f"strategic-ownership-matrix-{DATE}.json"
    xlsx_path = PLANS / f"مصفوفة_ملكية_الخطة_والمؤشرات_والأقسام_{DATE}.xlsx"

    write_markdown(md_path)
    write_json(json_path)
    write_excel(xlsx_path)

    print(f"Markdown: {md_path}")
    print(f"JSON: {json_path}")
    print(f"Excel: {xlsx_path}")
    print(f"Rows: {len(ROWS)}")


if __name__ == "__main__":
    main()
