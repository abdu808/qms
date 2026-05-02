from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
PLANS = ROOT / "ISO9001" / "الخطط والمشرات"
DATE = "2026-05-02"


OPERATING_RULES = [
    {
        "step": "تجهيز القيمة",
        "owner": "مالك البيانات",
        "when": "اليوم 1-5 من كل شهر",
        "action": "استخراج قيمة المؤشر من مصدره المعتمد، مثل نظام الجودة أو رافد أو السجلات المالية أو سجلات الوثائق.",
        "output": "قيمة رقمية أو نسبة مئوية جاهزة للإدخال.",
    },
    {
        "step": "إدخال القراءة",
        "owner": "مالك البيانات",
        "when": "اليوم 5-7 من كل شهر",
        "action": "إدخال قراءة المؤشر في النظام مع فترة القياس، وإضافة ملاحظة مختصرة إذا كانت القيمة غير مكتملة أو تحتاج تفسير.",
        "output": "قراءة شهرية موثقة في النظام.",
    },
    {
        "step": "مراجعة أولية",
        "owner": "مالك الأداء",
        "when": "اليوم 8-10 من كل شهر",
        "action": "مراجعة القراءة والتأكد من منطقيتها مقارنة بالمستهدف وخط الأساس والاتجاه.",
        "output": "قراءة مقبولة أو ملاحظة تصحيح للمالك البيانات.",
    },
    {
        "step": "تحليل الانحراف",
        "owner": "مالك الأداء",
        "when": "اليوم 10-12 من كل شهر",
        "action": "إذا ظهر انحراف أصفر أو أحمر، يكتب السبب المختصر والإجراء المقترح وموعد المعالجة.",
        "output": "سبب انحراف وإجراء تصحيحي مختصر عند الحاجة.",
    },
    {
        "step": "اعتماد القراءة",
        "owner": "جهة الاعتماد",
        "when": "اليوم 12-15 من كل شهر",
        "action": "اعتماد القراءة أو إعادتها للتصحيح إذا كانت غير موثقة أو غير منطقية.",
        "output": "قراءة معتمدة أو معادة للتصحيح.",
    },
    {
        "step": "تقرير شهري مختصر",
        "owner": "وحدة الاستراتيجية والتميز المؤسسي",
        "when": "اليوم 16-20 من كل شهر",
        "action": "تجميع المؤشرات المتأخرة أو المنحرفة ورفع ملخص قصير للمدير التنفيذي.",
        "output": "ملخص شهري من صفحة واحدة.",
    },
    {
        "step": "مراجعة ربع سنوية",
        "owner": "المدير التنفيذي + وحدة الاستراتيجية",
        "when": "نهاية كل ربع",
        "action": "مراجعة الاتجاهات والانحرافات المتكررة وربطها بالقرارات أو الإجراءات التصحيحية أو فرص التحسين.",
        "output": "محضر مراجعة أداء ربع سنوي.",
    },
]


DEVIATION_RULES = [
    {
        "status": "أخضر",
        "condition": "تحقق 95% أو أكثر من المستهدف، أو حسب عتبة المؤشر.",
        "required_action": "لا يلزم إجراء تصحيحي. تضاف ملاحظة اختيارية إذا كانت هناك فرصة تحسين.",
    },
    {
        "status": "أصفر",
        "condition": "تحقق بين 75% و94% من المستهدف، أو انحراف متوسط لا يهدد الهدف.",
        "required_action": "يكتب مالك الأداء سبب مختصر وإجراء متابعة بسيط بموعد واضح.",
    },
    {
        "status": "أحمر",
        "condition": "أقل من 75% من المستهدف، أو انحراف متكرر، أو خطر على التزام ISO أو هدف استراتيجي.",
        "required_action": "فتح إجراء تصحيحي أو مهمة متابعة، وتحديد مسؤول وتاريخ إغلاق.",
    },
    {
        "status": "رمادي",
        "condition": "لا توجد قراءة أو مصدر البيانات غير جاهز.",
        "required_action": "تحديد سبب عدم توفر القراءة وموعد توفيرها. إذا تكرر شهرين يرفع للمدير التنفيذي.",
    },
]


OWNER_CADENCE = [
    {
        "domain": "مؤشرات رافد والمستفيدين",
        "examples": "الكفالة، السلال، التمكين، رضا المستفيد، الأثر",
        "data_source": "رافد + ملخصات معتمدة",
        "data_owner": "إدارة الخدمة المجتمعية / المساعدات العينية حسب المؤشر",
        "note": "لا تنسخ التفاصيل من رافد؛ أدخل الملخص الشهري فقط في نظام الجودة.",
    },
    {
        "domain": "مؤشرات تنمية الموارد",
        "examples": "الإيرادات، الحملات، المانحون، استبقاء الداعمين",
        "data_source": "سجلات التبرعات والحملات + المالية للمطابقة",
        "data_owner": "إدارة تنمية الموارد والمشاريع",
        "note": "تثبت المالية الرقم عند الحاجة، لكنها لا تملك علاقة المانح.",
    },
    {
        "domain": "مؤشرات الاستثمار والمالية",
        "examples": "التعادل المالي، تغطية المصروفات، المحفظة، الإقفالات",
        "data_source": "السجلات المالية والتقارير المعتمدة",
        "data_owner": "الإدارة المالية / وحدة الاستثمار",
        "note": "الأرقام المالية تحتاج مطابقة قبل الاعتماد.",
    },
    {
        "domain": "مؤشرات الجودة وISO",
        "examples": "جاهزية ISO، الوثائق، التدقيق، الإجراءات التصحيحية",
        "data_source": "نظام الجودة وسجلات الوثائق والمراجعة",
        "data_owner": "وحدة الاستراتيجية والتميز المؤسسي",
        "note": "هذه هي مؤشرات نظام الجودة الأصلية ولا تعتمد على رافد.",
    },
    {
        "domain": "مؤشرات التقنية والتحول الرقمي",
        "examples": "توافر النظام، AI، الأتمتة",
        "data_source": "سجلات النظام وتقارير التقنية",
        "data_owner": "قسم تقنية المعلومات",
        "note": "تراجع وحدة الاستراتيجية المنهجية، والتقنية تملك التنفيذ والبيانات.",
    },
    {
        "domain": "مؤشرات الموارد البشرية والتطوع",
        "examples": "تدريب الموظفين، رضا الموظفين، مراجعات الأداء، ساعات التطوع",
        "data_source": "سجلات الموارد البشرية والتطوع",
        "data_owner": "قسم الموارد البشرية / الاتصال المؤسسي والشراكات",
        "note": "افصل تدريب الموظفين عن تدريب المستفيدين.",
    },
]


DEVIATION_TEMPLATE = [
    "كود المؤشر",
    "اسم المؤشر",
    "الشهر",
    "القيمة الفعلية",
    "المستهدف",
    "الحالة",
    "سبب الانحراف",
    "الإجراء المقترح",
    "المسؤول",
    "تاريخ الإغلاق المتوقع",
    "حالة الإجراء",
]


def write_markdown(path: Path) -> None:
    lines = [
        "# دليل التشغيل الشهري للمؤشرات وتصحيح الانحرافات",
        "",
        f"**التاريخ:** {DATE}",
        "",
        "## الغرض",
        "",
        "تسهيل إدخال قيم المؤشرات ومراجعتها وتصحيح الانحرافات بدون تعقيد، مع بقاء رافد مصدرًا تفصيليًا لبرامج المستفيدين، ونظام الجودة مصدرًا للمتابعة والحوكمة والتقارير.",
        "",
        "## دورة العمل الشهرية",
        "",
        "| الخطوة | المسؤول | التوقيت | ماذا يعمل؟ | الناتج |",
        "|---|---|---|---|---|",
    ]
    for row in OPERATING_RULES:
        lines.append(f"| {row['step']} | {row['owner']} | {row['when']} | {row['action']} | {row['output']} |")

    lines.extend(["", "## قواعد التعامل مع الانحرافات", "", "| الحالة | متى؟ | المطلوب |", "|---|---|---|"])
    for row in DEVIATION_RULES:
        lines.append(f"| {row['status']} | {row['condition']} | {row['required_action']} |")

    lines.extend(["", "## مصادر البيانات حسب المجال", "", "| المجال | أمثلة | مصدر البيانات | مالك البيانات | ملاحظة |", "|---|---|---|---|---|"])
    for row in OWNER_CADENCE:
        lines.append(
            f"| {row['domain']} | {row['examples']} | {row['data_source']} | {row['data_owner']} | {row['note']} |"
        )

    lines.extend(
        [
            "",
            "## قالب سبب الانحراف",
            "",
            "عند ظهور حالة صفراء أو حمراء، تكفي صياغة قصيرة بهذه الطريقة:",
            "",
            "> السبب: تأخر تحديث البيانات من المصدر / نقص مستند / تأخر تنفيذ / اعتماد لم يكتمل.",
            "> الإجراء: تحديث السجل، مخاطبة المالك، فتح إجراء تصحيحي، أو تعديل خطة تنفيذ.",
            "",
            "## قاعدة عدم التعقيد",
            "",
            "- لا يكتب تقرير طويل لكل مؤشر.",
            "- لا تكرر بيانات رافد داخل نظام الجودة.",
            "- لا تفتح إجراء تصحيحي لكل انحراف بسيط.",
            "- الإجراء التصحيحي يفتح فقط عند الانحراف الأحمر، أو الانحراف المتكرر، أو ما يؤثر على ISO أو هدف استراتيجي.",
            "- الملخص الشهري يجب أن يكون صفحة واحدة قدر الإمكان.",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def write_json(path: Path) -> None:
    payload = {
        "title": "دليل التشغيل الشهري للمؤشرات وتصحيح الانحرافات",
        "date": DATE,
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "operatingRules": OPERATING_RULES,
        "deviationRules": DEVIATION_RULES,
        "ownerCadence": OWNER_CADENCE,
        "deviationTemplateColumns": DEVIATION_TEMPLATE,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def style_sheet(ws) -> None:
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
            cell.border = border
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font


def autosize(ws, widths: dict[int, int] | None = None) -> None:
    widths = widths or {}
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        if col_idx in widths:
            ws.column_dimensions[letter].width = widths[col_idx]
            continue
        width = 12
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row_idx, col_idx).value
            if value:
                width = max(width, min(55, len(str(value)) + 3))
        ws.column_dimensions[letter].width = width


def write_excel(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "دورة التشغيل"
    ws.append(["الخطوة", "المسؤول", "التوقيت", "ماذا يعمل؟", "الناتج"])
    for row in OPERATING_RULES:
        ws.append([row["step"], row["owner"], row["when"], row["action"], row["output"]])
    style_sheet(ws)
    autosize(ws, {1: 22, 2: 28, 3: 24, 4: 70, 5: 40})

    ws2 = wb.create_sheet("قواعد الانحراف")
    ws2.append(["الحالة", "متى؟", "المطلوب"])
    for row in DEVIATION_RULES:
        ws2.append([row["status"], row["condition"], row["required_action"]])
    style_sheet(ws2)
    autosize(ws2, {1: 14, 2: 65, 3: 75})

    ws3 = wb.create_sheet("مصادر البيانات")
    ws3.append(["المجال", "أمثلة", "مصدر البيانات", "مالك البيانات", "ملاحظة"])
    for row in OWNER_CADENCE:
        ws3.append([row["domain"], row["examples"], row["data_source"], row["data_owner"], row["note"]])
    style_sheet(ws3)
    autosize(ws3, {1: 32, 2: 45, 3: 45, 4: 42, 5: 70})

    ws4 = wb.create_sheet("قالب الانحرافات")
    ws4.append(DEVIATION_TEMPLATE)
    for _ in range(12):
        ws4.append([""] * len(DEVIATION_TEMPLATE))
    style_sheet(ws4)
    autosize(ws4, {1: 16, 2: 32, 3: 14, 4: 16, 5: 16, 6: 14, 7: 45, 8: 45, 9: 24, 10: 24, 11: 18})

    wb.save(path)


def main() -> None:
    DOCS.mkdir(parents=True, exist_ok=True)
    PLANS.mkdir(parents=True, exist_ok=True)

    md_path = DOCS / f"monthly-kpi-operating-guide-{DATE}.md"
    json_path = DOCS / f"monthly-kpi-operating-guide-{DATE}.json"
    xlsx_path = PLANS / f"دليل_التشغيل_الشهري_للمؤشرات_وتصحيح_الانحرافات_{DATE}.xlsx"

    write_markdown(md_path)
    write_json(json_path)
    write_excel(xlsx_path)

    print(f"Markdown: {md_path}")
    print(f"JSON: {json_path}")
    print(f"Excel: {xlsx_path}")
    print("Operating steps:", len(OPERATING_RULES))
    print("Deviation rules:", len(DEVIATION_RULES))


if __name__ == "__main__":
    main()
