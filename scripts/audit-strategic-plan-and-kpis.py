import json
import os
import re
import sys
import urllib.request
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TODAY = "2026-05-02"
PLAN_DIR = ROOT / "ISO9001" / "الخطط والمشرات"
OUT_JSON = ROOT / "docs" / f"strategic-plan-kpi-audit-{TODAY}.json"
OUT_MD = ROOT / "docs" / f"strategic-plan-kpi-audit-{TODAY}.md"

FILES = {
    "strategic_plan": PLAN_DIR / "الخطة_الاستراتيجية_2026_2030_المعتمدة.xlsx",
    "kpi_catalog": PLAN_DIR / "KPI_Catalog_v11_2026_2030.xlsx",
    "annual_targets": PLAN_DIR / "AnnualTargets_2026_2030.xlsx",
    "operational_plan": PLAN_DIR / "الخطة_التشغيلية_2026_بـ4_محاور.xlsx",
    "initiatives": PLAN_DIR / "المبادرات_الـ21_2026_2030.xlsx",
}


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("%", "")
    if text in {"", "—", "-"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def find_header_row(ws, required):
    required = set(required)
    for row_idx in range(1, min(ws.max_row, 15) + 1):
        values = [clean(c.value) for c in ws[row_idx]]
        if required.issubset(set(v for v in values if v is not None)):
            return row_idx, values
    raise ValueError(f"Could not find header row in {ws.title}: {required}")


def table_from_sheet(path, sheet_name=None, required=("الكود",)):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    header_row, headers = find_header_row(ws, required)
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        item = {}
        for idx, header in enumerate(headers):
            if header is None:
                continue
            item[str(header)] = clean(row[idx] if idx < len(row) else None)
        if any(v is not None for v in item.values()):
            rows.append(item)
    return rows


def api_get(base_url, token, path):
    req = urllib.request.Request(
        base_url.rstrip("/") + path,
        headers={
            "authorization": f"Bearer {token}",
            "content-type": "application/json",
            "user-agent": "node",
        },
    )
    with urllib.request.urlopen(req, timeout=45) as resp:
        return json.loads(resp.read().decode("utf-8"))


def api_login(base_url, email, password):
    body = json.dumps({"email": email, "password": password}).encode("utf-8")
    req = urllib.request.Request(
        base_url.rstrip("/") + "/api/auth/login",
        data=body,
        method="POST",
        headers={"content-type": "application/json", "user-agent": "node"},
    )
    with urllib.request.urlopen(req, timeout=45) as resp:
        data = json.loads(resp.read().decode("utf-8"))
    return data["token"]


def code_set(rows, field):
    return {str(r.get(field)).strip() for r in rows if r.get(field)}


def duplicate_pairs(rows, fields):
    counts = Counter(
        tuple(str(r.get(field)).strip() for field in fields)
        for r in rows
        if all(r.get(field) is not None for field in fields)
    )
    return sorted([pair for pair, count in counts.items() if count > 1])


def duplicates(rows, field):
    counts = Counter(str(r.get(field)).strip() for r in rows if r.get(field))
    return sorted([code for code, count in counts.items() if count > 1])


def pct_sum(rows, field):
    return round(sum(number(r.get(field)) or 0 for r in rows), 2)


def extract_axis_code(value):
    if not value:
        return None
    match = re.search(r"AXIS-\d+", str(value))
    return match.group(0) if match else None


def severity_score(level):
    return {"حرج": 3, "عال": 2, "متوسط": 1, "منخفض": 0}.get(level, 0)


def add_finding(findings, level, area, finding, recommendation, evidence=None):
    findings.append(
        {
            "severity": level,
            "area": area,
            "finding": finding,
            "recommendation": recommendation,
            "evidence": evidence or "",
        }
    )


def md_table(rows, headers):
    if not rows:
        return ""
    out = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        out.append("| " + " | ".join(str(row.get(h, "")).replace("\n", "<br>") for h in headers) + " |")
    return "\n".join(out)


def main():
    for name, path in FILES.items():
        if not path.exists():
            raise FileNotFoundError(f"Missing {name}: {path}")

    goals = table_from_sheet(FILES["strategic_plan"], "الخطة الاستراتيجية", required=("كود الهدف",))
    axes = table_from_sheet(FILES["strategic_plan"], "المحاور الاستراتيجية", required=("الكود", "اسم المحور"))
    indicators = table_from_sheet(FILES["kpi_catalog"], required=("الكود", "اسم المؤشر"))
    annual_targets = table_from_sheet(FILES["annual_targets"], required=("كود المؤشر", "السنة"))
    objectives = table_from_sheet(FILES["operational_plan"], required=("الكود", "الهدف الاستراتيجي"))
    initiatives = table_from_sheet(FILES["initiatives"], required=("الكود", "الهدف الاستراتيجي"))
    goals = [r for r in goals if r.get("كود الهدف")]
    axes = [r for r in axes if r.get("الكود")]
    indicators = [r for r in indicators if r.get("الكود")]
    annual_targets = [r for r in annual_targets if r.get("كود المؤشر")]
    objectives = [r for r in objectives if r.get("الكود")]
    initiatives = [r for r in initiatives if r.get("الكود")]

    findings = []
    goal_codes = code_set(goals, "كود الهدف")
    axis_codes = code_set(axes, "الكود")
    indicator_codes = code_set(indicators, "الكود")
    objective_codes = code_set(objectives, "الكود")
    initiative_codes = code_set(initiatives, "الكود")

    for title, rows, field in [
        ("الأهداف الاستراتيجية", goals, "كود الهدف"),
        ("المحاور", axes, "الكود"),
        ("المؤشرات", indicators, "الكود"),
        ("الأهداف التشغيلية", objectives, "الكود"),
        ("المبادرات", initiatives, "الكود"),
    ]:
        dup = duplicates(rows, field)
        if dup:
            add_finding(findings, "حرج", title, f"توجد أكواد مكررة: {', '.join(dup)}", "توحيد الأكواد قبل أي استيراد أو اعتماد.")

    annual_dup = duplicate_pairs(annual_targets, ("كود المؤشر", "السنة"))
    if annual_dup:
        sample = ", ".join(f"{code}/{year}" for code, year in annual_dup[:20])
        add_finding(findings, "حرج", "المستهدفات السنوية", f"توجد مستهدفات مكررة لنفس المؤشر والسنة: {sample}", "الإبقاء على سجل واحد لكل مؤشر/سنة.")

    axis_weight_sum = pct_sum(axes, "الوزن %")
    if axis_weight_sum != 100:
        add_finding(findings, "عال", "المحاور", f"مجموع أوزان المحاور = {axis_weight_sum}% وليس 100%.", "مراجعة أوزان المحاور حتى تساوي 100%.")

    indicator_weight_sum = pct_sum(indicators, "الوزن %")
    if indicator_weight_sum != 100:
        add_finding(findings, "عال", "المؤشرات", f"مجموع أوزان المؤشرات = {indicator_weight_sum}% وليس 100%.", "تعديل أوزان المؤشرات قبل ربطها بلوحات الأداء.")

    goal_axis_counts = Counter()
    for goal in goals:
        axis_code = extract_axis_code(goal.get("المحور"))
        if not axis_code or axis_code not in axis_codes:
            add_finding(findings, "عال", "ربط الأهداف بالمحاور", f"الهدف {goal.get('كود الهدف')} مرتبط بمحور غير معروف: {goal.get('المحور')}", "اختيار محور موجود من جدول المحاور.")
        else:
            goal_axis_counts[axis_code] += 1

    for axis in axes:
        declared = int(number(axis.get("عدد الأهداف")) or 0)
        actual = goal_axis_counts.get(axis.get("الكود"), 0)
        if declared != actual:
            add_finding(findings, "متوسط", "المحاور", f"{axis.get('الكود')} عدد الأهداف المعلن {declared} بينما الفعلي {actual}.", "تحديث حقل عدد الأهداف أو مراجعة ربط الأهداف.")

    bad_objectives = [r.get("الكود") for r in objectives if r.get("الهدف الاستراتيجي") not in goal_codes]
    if bad_objectives:
        add_finding(findings, "حرج", "الأهداف التشغيلية", f"{len(bad_objectives)} أهداف تشغيلية مرتبطة بأهداف استراتيجية غير موجودة.", "تصحيح حقل الهدف الاستراتيجي قبل الاستيراد.", ", ".join(bad_objectives[:20]))

    bad_initiatives = [r.get("الكود") for r in initiatives if r.get("الهدف الاستراتيجي") not in goal_codes]
    if bad_initiatives:
        add_finding(findings, "حرج", "المبادرات", f"{len(bad_initiatives)} مبادرات مرتبطة بأهداف استراتيجية غير موجودة.", "تصحيح حقل الهدف الاستراتيجي قبل الاستيراد.", ", ".join(bad_initiatives[:20]))

    target_by_indicator = defaultdict(list)
    for row in annual_targets:
        target_by_indicator[row.get("كود المؤشر")].append(row)

    missing_target_indicators = sorted(indicator_codes - set(target_by_indicator))
    if missing_target_indicators:
        add_finding(findings, "حرج", "المستهدفات السنوية", "توجد مؤشرات بلا مستهدفات سنوية.", "إضافة مستهدفات 2026-2030 لكل مؤشر.", ", ".join(missing_target_indicators))

    orphan_targets = sorted(set(target_by_indicator) - indicator_codes)
    if orphan_targets:
        add_finding(findings, "عال", "المستهدفات السنوية", "توجد مستهدفات لمؤشرات غير موجودة في الكتالوج.", "حذفها أو إضافة المؤشر الناقص في الكتالوج.", ", ".join(orphan_targets))

    incomplete_targets = []
    expected_years = {2026, 2027, 2028, 2029, 2030}
    for code in indicator_codes:
        years = {int(number(r.get("السنة")) or 0) for r in target_by_indicator.get(code, [])}
        if years != expected_years:
            incomplete_targets.append(f"{code}: {sorted(years)}")
    if incomplete_targets:
        add_finding(findings, "عال", "المستهدفات السنوية", "ليست كل المؤشرات لديها السنوات الخمس كاملة.", "استكمال السنوات 2026-2030 لكل مؤشر.", "; ".join(incomplete_targets[:20]))

    indicators_without_baseline = [r.get("الكود") for r in indicators if r.get("خط الأساس") in (None, "—", "-")]
    if indicators_without_baseline:
        add_finding(findings, "متوسط", "المؤشرات", f"{len(indicators_without_baseline)} مؤشرات بلا خط أساس رقمي.", "تعبئة خط الأساس أو توثيق سبب عدم توفره ومتى سيقاس.", ", ".join(indicators_without_baseline))

    missing_owners = []
    for source, rows, code_field, owner_field in [
        ("الأهداف الاستراتيجية", goals, "كود الهدف", "المالك"),
        ("المؤشرات", indicators, "الكود", "المالك"),
        ("المبادرات", initiatives, "الكود", "المالك"),
    ]:
        for row in rows:
            if not row.get(owner_field):
                missing_owners.append(f"{source}: {row.get(code_field)}")
    if missing_owners:
        add_finding(findings, "عال", "الملكية", "توجد عناصر بلا مالك واضح.", "تعيين مالك بالاسم لكل عنصر قبل التفعيل.", ", ".join(missing_owners[:30]))

    if "كود المؤشر" not in objectives[0] if objectives else True:
        add_finding(
            findings,
            "عال",
            "ربط المؤشرات بالأهداف التشغيلية",
            "الخطة التشغيلية تحتوي وصف KPI نصي، ولا تحتوي كود مؤشر معتمد يربطها بكتالوج المؤشرات.",
            "إضافة عمود indicatorCode لكل هدف تشغيلي أو اعتماد خريطة ربط رسمية OBJ ↔ IND.",
        )

    if any("13,000" in str(r.get("الهدف التشغيلي")) or "13000" in str(r.get("المستهدف")) for r in objectives):
        add_finding(
            findings,
            "عال",
            "واقعية المستهدفات",
            "هدف الوصول إلى 13,000 أسرة يحتاج تعريفاً حاكماً لأنه يتجاوز عدد ملفات المستفيدين الفريدة المعروفة سابقاً.",
            "تحديد هل المقصود أسر فريدة أم خدمات تراكمية. إن كان خدمات تراكمية فيكتب ذلك صراحة في اسم المؤشر ووحدته.",
            "خط الأساس السابق يشير إلى 2,375 ملف مستفيد فعلي.",
        )

    if any("870" in str(r.get("اسم المبادرة")) and "يتيم" in str(r.get("اسم المبادرة")) for r in initiatives):
        add_finding(
            findings,
            "عال",
            "واقعية المبادرات",
            "مبادرة كفالة 870 يتيم تحتاج إعادة ضبط أو تعريف، لأن خط الأساس السابق فرّق بين إجمالي الأيتام واحتياج الكفالة الفعلي.",
            "تعديل المبادرة إلى نطاق قابل للتنفيذ مثل كفالة 200 يتيم بحلول 2027 أو اعتماد تعريف: مستفيدون متأثرون/مكفولون.",
        )

    live = {"available": False, "errors": []}
    base_url = os.environ.get("QMS_BASE_URL")
    email = os.environ.get("QMS_EMAIL")
    password = os.environ.get("QMS_PASSWORD")
    if base_url and email and password:
        try:
            token = api_login(base_url, email, password)
            endpoints = {
                "plans": "/api/strategic-plans?limit=100",
                "goals": "/api/strategic-goals?limit=100",
                "objectives": "/api/objectives?limit=100",
                "indicators": "/api/indicators?limit=100",
                "initiatives": "/api/initiatives?limit=100",
                "indicator_weight_check": "/api/indicators/weight-check",
                "kpi_dashboard": "/api/kpi/dashboard?year=2026&month=5",
            }
            data = {}
            for key, path in endpoints.items():
                try:
                    data[key] = api_get(base_url, token, path)
                except Exception as exc:
                    live["errors"].append(f"{key}: {exc}")
            live["available"] = True
            live["baseUrl"] = base_url
            live["counts"] = {
                key: data.get(key, {}).get("total")
                for key in ["plans", "goals", "objectives", "indicators", "initiatives"]
            }
            live["indicatorWeightTotal"] = data.get("indicator_weight_check", {}).get("total")
            live["indicatorWeightValid"] = data.get("indicator_weight_check", {}).get("isValid")
            live["dashboardSummary"] = data.get("kpi_dashboard", {}).get("summary")

            api_goal_codes = code_set(data.get("goals", {}).get("items", []), "code")
            api_objective_codes = code_set(data.get("objectives", {}).get("items", []), "code")
            api_indicator_codes = code_set(data.get("indicators", {}).get("items", []), "code")
            api_initiative_codes = code_set(data.get("initiatives", {}).get("items", []), "code")

            comparisons = {
                "goals_missing_in_api": sorted(goal_codes - api_goal_codes),
                "goals_extra_in_api": sorted(api_goal_codes - goal_codes),
                "objectives_missing_in_api": sorted(objective_codes - api_objective_codes),
                "objectives_extra_in_api": sorted(api_objective_codes - objective_codes),
                "indicators_missing_in_api": sorted(indicator_codes - api_indicator_codes),
                "indicators_extra_in_api": sorted(api_indicator_codes - indicator_codes),
                "initiatives_missing_in_api": sorted(initiative_codes - api_initiative_codes),
                "initiatives_extra_in_api": sorted(api_initiative_codes - initiative_codes),
            }
            live["comparisons"] = comparisons

            for label, missing in [
                ("الأهداف الاستراتيجية", comparisons["goals_missing_in_api"]),
                ("الأهداف التشغيلية", comparisons["objectives_missing_in_api"]),
                ("المؤشرات", comparisons["indicators_missing_in_api"]),
                ("المبادرات", comparisons["initiatives_missing_in_api"]),
            ]:
                if missing:
                    add_finding(findings, "عال", "مطابقة النظام الحي", f"{label}: عناصر موجودة في الملف وغير موجودة في النظام الحي.", "استكمال الاستيراد أو تحديث الملف المعتمد.", ", ".join(missing[:30]))

            for label, extra in [
                ("الأهداف الاستراتيجية", comparisons["goals_extra_in_api"]),
                ("الأهداف التشغيلية", comparisons["objectives_extra_in_api"]),
                ("المؤشرات", comparisons["indicators_extra_in_api"]),
                ("المبادرات", comparisons["initiatives_extra_in_api"]),
            ]:
                if extra:
                    add_finding(findings, "متوسط", "مطابقة النظام الحي", f"{label}: عناصر موجودة في النظام الحي وغير موجودة في الملف.", "تحديد هل هي قديمة فتؤرشف، أو صحيحة فتضاف للملف.", ", ".join(extra[:30]))

            if live.get("indicatorWeightValid") is False:
                add_finding(findings, "عال", "النظام الحي", f"مجموع أوزان المؤشرات في النظام الحي = {live.get('indicatorWeightTotal')} وليس 100.", "تعديل أوزان المؤشرات في النظام قبل اعتماد لوحة الأداء.")
        except Exception as exc:
            live["errors"].append(str(exc))

    findings.sort(key=lambda f: severity_score(f["severity"]), reverse=True)

    summary = {
        "files": {key: str(path.relative_to(ROOT)).replace("\\", "/") for key, path in FILES.items()},
        "counts": {
            "axes": len(axes),
            "strategicGoals": len(goals),
            "operationalObjectives": len(objectives),
            "indicators": len(indicators),
            "annualTargets": len(annual_targets),
            "initiatives": len(initiatives),
        },
        "weights": {
            "axisWeightSum": axis_weight_sum,
            "indicatorWeightSum": indicator_weight_sum,
        },
        "live": live,
        "findings": findings,
    }

    OUT_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    critical = sum(1 for f in findings if f["severity"] == "حرج")
    high = sum(1 for f in findings if f["severity"] == "عال")
    medium = sum(1 for f in findings if f["severity"] == "متوسط")
    verdict = "قابلة للتطبيق بعد إغلاق الملاحظات العالية والحرجة" if critical or high else "قابلة للتطبيق مباشرة"

    finding_rows = [
        {
            "الأولوية": f["severity"],
            "المجال": f["area"],
            "الملاحظة": f["finding"],
            "الإجراء المقترح": f["recommendation"],
            "الدليل": f["evidence"],
        }
        for f in findings
    ]
    md = f"""# مراجعة الخطة الاستراتيجية والمؤشرات 2026-2030

**التاريخ:** {datetime.now().strftime("%Y-%m-%d %H:%M")}  
**القرار المهني:** {verdict}

## الملخص التنفيذي

- المحاور: {len(axes)}
- الأهداف الاستراتيجية: {len(goals)}
- الأهداف التشغيلية: {len(objectives)}
- المؤشرات: {len(indicators)}
- المستهدفات السنوية: {len(annual_targets)}
- المبادرات: {len(initiatives)}
- مجموع أوزان المحاور: {axis_weight_sum}%
- مجموع أوزان المؤشرات: {indicator_weight_sum}%
- الملاحظات الحرجة: {critical}
- الملاحظات العالية: {high}
- الملاحظات المتوسطة: {medium}

## قراءة النظام الحي

{json.dumps(live, ensure_ascii=False, indent=2)}

## الملاحظات والإجراءات

{md_table(finding_rows, ["الأولوية", "المجال", "الملاحظة", "الإجراء المقترح", "الدليل"])}

## الحكم العملي

الخطة ليست بعيدة عن الاعتماد، لكنها تحتاج ضبطاً قبل أن تصبح قابلة للتنفيذ اليومي داخل النظام. أهم نقطة ليست وجود المالكين فقط، بل وجود ربط آلي واضح بين الهدف التشغيلي والمؤشر المعتمد، وتثبيت تعريفات المستهدفات التي قد تُفهم بأكثر من معنى مثل الأسر الفريدة مقابل الخدمات التراكمية.

## توصية التنفيذ

1. اعتماد خريطة ربط `OBJ ↔ IND` وإضافتها للملف أو النظام.
2. إغلاق تضارب المستهدفات ذات الحساسية العالية: 13,000 أسرة، 870 يتيم.
3. استكمال خطوط الأساس غير الرقمية أو توثيق موعد قياسها الأول.
4. مطابقة الأكواد بين الملفات والنظام الحي، ثم تنفيذ دفعة تحديث واحدة.
5. بعد التحديث، تشغيل تقرير الوزن والترابط مرة أخرى قبل تجميد الخطة.
"""
    OUT_MD.write_text(md, encoding="utf-8")

    print(json.dumps({
        "ok": True,
        "report": str(OUT_MD.relative_to(ROOT)).replace("\\", "/"),
        "json": str(OUT_JSON.relative_to(ROOT)).replace("\\", "/"),
        "counts": summary["counts"],
        "weights": summary["weights"],
        "findings": {"critical": critical, "high": high, "medium": medium},
        "liveAvailable": live["available"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        raise
