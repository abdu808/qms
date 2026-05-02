from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
PLANS = ROOT / "ISO9001" / "الخطط والمشرات"
DATE = "2026-05-02"


ROWS = [
    {
        "program_code": "PRG-2026-001",
        "program": "برنامج ISO 9001 والجودة المؤسسية",
        "indicator_code": "QOP-2026-001",
        "indicator": "نسبة جاهزية متطلبات ISO 9001",
        "definition": "نسبة متطلبات ISO 9001 المكتملة والمعتمدة من إجمالي المتطلبات المحددة في قائمة الجاهزية.",
        "formula": "المتطلبات المكتملة ÷ إجمالي المتطلبات × 100",
        "unit": "%",
        "frequency": "شهري",
        "owner": "وحدة الاستراتيجية والتميز المؤسسي",
        "data_owner": "قسم الجودة والتميز المؤسسي",
        "approver": "المدير التنفيذي",
        "source": "نظام الجودة / قائمة جاهزية ISO",
        "target_2026": "100%",
        "priority": "عال",
        "recommendation": "يضاف للنظام",
    },
    {
        "program_code": "PRG-2026-001",
        "program": "برنامج ISO 9001 والجودة المؤسسية",
        "indicator_code": "QOP-2026-002",
        "indicator": "نسبة إغلاق ملاحظات التدقيق الخارجي",
        "definition": "نسبة ملاحظات التدقيق الخارجي المغلقة في موعدها من إجمالي الملاحظات المفتوحة.",
        "formula": "الملاحظات المغلقة في موعدها ÷ إجمالي الملاحظات × 100",
        "unit": "%",
        "frequency": "شهري أثناء فترة التدقيق",
        "owner": "وحدة الاستراتيجية والتميز المؤسسي",
        "data_owner": "قسم الجودة والتميز المؤسسي",
        "approver": "المدير التنفيذي",
        "source": "سجل الملاحظات / تقارير الجهة المانحة للشهادة",
        "target_2026": "100%",
        "priority": "عال",
        "recommendation": "يضاف للنظام",
    },
    {
        "program_code": "PRG-2026-002",
        "program": "برنامج حوكمة الخطة والمؤشرات",
        "indicator_code": "QOP-2026-003",
        "indicator": "نسبة تحديث المؤشرات في موعدها",
        "definition": "نسبة المؤشرات التي تم تحديث قراءاتها خلال الفترة المحددة من إجمالي المؤشرات المستحقة للتحديث.",
        "formula": "المؤشرات المحدثة في موعدها ÷ المؤشرات المستحقة × 100",
        "unit": "%",
        "frequency": "شهري",
        "owner": "وحدة الاستراتيجية والتميز المؤسسي",
        "data_owner": "جميع الإدارات حسب المؤشر",
        "approver": "وحدة الاستراتيجية والتميز المؤسسي",
        "source": "نظام الجودة / قراءات المؤشرات",
        "target_2026": "95%",
        "priority": "عال",
        "recommendation": "يضاف للنظام",
    },
    {
        "program_code": "PRG-2026-002",
        "program": "برنامج حوكمة الخطة والمؤشرات",
        "indicator_code": "QOP-2026-004",
        "indicator": "نسبة اكتمال تقارير المراجعة الربع سنوية",
        "definition": "نسبة تقارير مراجعة الأداء الربع سنوية المكتملة والمعتمدة من إجمالي التقارير المخطط لها.",
        "formula": "التقارير المكتملة والمعتمدة ÷ التقارير المخطط لها × 100",
        "unit": "%",
        "frequency": "ربع سنوي",
        "owner": "وحدة الاستراتيجية والتميز المؤسسي",
        "data_owner": "وحدة الاستراتيجية والتميز المؤسسي",
        "approver": "المدير التنفيذي",
        "source": "محاضر المراجعة / تقارير الأداء",
        "target_2026": "100%",
        "priority": "عال",
        "recommendation": "يضاف للنظام",
    },
    {
        "program_code": "PRG-2026-003",
        "program": "برنامج إدارة الوثائق والسياسات والإجراءات",
        "indicator_code": "QOP-2026-005",
        "indicator": "نسبة الوثائق والسياسات المعتمدة",
        "definition": "نسبة الوثائق والسياسات والإجراءات المعتمدة من إجمالي الوثائق المطلوب اعتمادها ضمن نطاق نظام الجودة.",
        "formula": "الوثائق المعتمدة ÷ الوثائق المطلوبة × 100",
        "unit": "%",
        "frequency": "شهري",
        "owner": "وحدة الاستراتيجية والتميز المؤسسي",
        "data_owner": "قسم الجودة والتميز المؤسسي",
        "approver": "المدير التنفيذي",
        "source": "سجل الوثائق والسياسات",
        "target_2026": "100%",
        "priority": "عال",
        "recommendation": "يضاف للنظام",
    },
    {
        "program_code": "PRG-2026-003",
        "program": "برنامج إدارة الوثائق والسياسات والإجراءات",
        "indicator_code": "QOP-2026-006",
        "indicator": "متوسط زمن اعتماد الوثيقة",
        "definition": "متوسط عدد الأيام من تاريخ رفع الوثيقة للمراجعة إلى تاريخ اعتمادها النهائي.",
        "formula": "مجموع أيام الاعتماد ÷ عدد الوثائق المعتمدة",
        "unit": "يوم",
        "frequency": "شهري",
        "owner": "وحدة الاستراتيجية والتميز المؤسسي",
        "data_owner": "قسم الجودة والتميز المؤسسي",
        "approver": "المدير التنفيذي",
        "source": "سجل الوثائق / تواريخ الاعتماد",
        "target_2026": "≤ 10 أيام",
        "priority": "متوسط",
        "recommendation": "مرحلة ثانية",
    },
    {
        "program_code": "PRG-2026-004",
        "program": "برنامج المراجعة الداخلية وإغلاق الملاحظات",
        "indicator_code": "QOP-2026-007",
        "indicator": "نسبة تنفيذ خطة المراجعة الداخلية",
        "definition": "نسبة مهام/زيارات المراجعة الداخلية المنفذة من إجمالي ما هو مخطط في خطة المراجعة.",
        "formula": "المراجعات المنفذة ÷ المراجعات المخططة × 100",
        "unit": "%",
        "frequency": "ربع سنوي",
        "owner": "وحدة الرقابة الداخلية",
        "data_owner": "وحدة الرقابة الداخلية",
        "approver": "لجنة المراجعة",
        "source": "خطة المراجعة الداخلية",
        "target_2026": "100%",
        "priority": "عال",
        "recommendation": "يضاف للنظام",
    },
    {
        "program_code": "PRG-2026-004",
        "program": "برنامج المراجعة الداخلية وإغلاق الملاحظات",
        "indicator_code": "QOP-2026-008",
        "indicator": "نسبة إغلاق ملاحظات الرقابة الداخلية في موعدها",
        "definition": "نسبة ملاحظات الرقابة الداخلية التي أغلقت ضمن المدة المحددة من إجمالي الملاحظات المستحقة للإغلاق.",
        "formula": "الملاحظات المغلقة في موعدها ÷ الملاحظات المستحقة × 100",
        "unit": "%",
        "frequency": "شهري",
        "owner": "وحدة الرقابة الداخلية",
        "data_owner": "وحدة الرقابة الداخلية",
        "approver": "لجنة المراجعة",
        "source": "سجل الملاحظات والإجراءات التصحيحية",
        "target_2026": "90%",
        "priority": "عال",
        "recommendation": "يضاف للنظام",
    },
    {
        "program_code": "PRG-2026-005",
        "program": "برنامج قياس الرضا وتجربة أصحاب المصلحة",
        "indicator_code": "QOP-2026-009",
        "indicator": "نسبة تنفيذ استبيانات الرضا المخططة",
        "definition": "نسبة استبيانات الرضا المنفذة خلال الفترة من إجمالي الاستبيانات المخطط لها لأصحاب المصلحة.",
        "formula": "الاستبيانات المنفذة ÷ الاستبيانات المخططة × 100",
        "unit": "%",
        "frequency": "ربع سنوي",
        "owner": "وحدة الاستراتيجية والتميز المؤسسي",
        "data_owner": "وحدة الاستراتيجية والتميز المؤسسي",
        "approver": "المدير التنفيذي",
        "source": "نظام الاستبيانات / نظام الجودة",
        "target_2026": "100%",
        "priority": "عال",
        "recommendation": "يضاف للنظام",
    },
    {
        "program_code": "PRG-2026-005",
        "program": "برنامج قياس الرضا وتجربة أصحاب المصلحة",
        "indicator_code": "QOP-2026-010",
        "indicator": "متوسط رضا أصحاب المصلحة",
        "definition": "متوسط نتائج الرضا للمستفيدين والمانحين والموظفين والشركاء حسب الاستبيانات المعتمدة.",
        "formula": "مجموع درجات الرضا ÷ عدد الاستجابات",
        "unit": "%",
        "frequency": "ربع سنوي",
        "owner": "وحدة الاستراتيجية والتميز المؤسسي",
        "data_owner": "الإدارات المالكة لكل فئة",
        "approver": "المدير التنفيذي",
        "source": "نتائج الاستبيانات",
        "target_2026": "تحديد خط أساس 2026 ثم تحسين سنوي",
        "priority": "متوسط",
        "recommendation": "مرحلة ثانية",
    },
    {
        "program_code": "PRG-2026-006",
        "program": "برنامج التحسين المستمر والإجراءات التصحيحية",
        "indicator_code": "QOP-2026-011",
        "indicator": "نسبة إغلاق الإجراءات التصحيحية في موعدها",
        "definition": "نسبة الإجراءات التصحيحية والوقائية المغلقة في الموعد من إجمالي الإجراءات المستحقة.",
        "formula": "الإجراءات المغلقة في موعدها ÷ الإجراءات المستحقة × 100",
        "unit": "%",
        "frequency": "شهري",
        "owner": "وحدة الاستراتيجية والتميز المؤسسي",
        "data_owner": "الإدارات المالكة للإجراء",
        "approver": "وحدة الاستراتيجية والتميز المؤسسي",
        "source": "سجل CAPA / فرص التحسين",
        "target_2026": "90%",
        "priority": "عال",
        "recommendation": "يضاف للنظام",
    },
    {
        "program_code": "PRG-2026-006",
        "program": "برنامج التحسين المستمر والإجراءات التصحيحية",
        "indicator_code": "QOP-2026-012",
        "indicator": "عدد فرص التحسين المنجزة",
        "definition": "عدد فرص التحسين التي تم تنفيذها وتوثيق أثرها خلال الفترة.",
        "formula": "عدد فرص التحسين المنجزة والمعتمدة",
        "unit": "عدد",
        "frequency": "ربع سنوي",
        "owner": "وحدة الاستراتيجية والتميز المؤسسي",
        "data_owner": "جميع الإدارات",
        "approver": "المدير التنفيذي",
        "source": "سجل التحسين المستمر",
        "target_2026": "8 فرص تحسين موثقة",
        "priority": "متوسط",
        "recommendation": "مرحلة ثانية",
    },
    {
        "program_code": "PRG-2026-007",
        "program": "برنامج التحول الرقمي ودعم نظام الجودة",
        "indicator_code": "QOP-2026-013",
        "indicator": "توافر نظام الجودة",
        "definition": "نسبة توافر نظام الجودة خلال الفترة مقارنة بإجمالي ساعات التشغيل المستهدفة.",
        "formula": "ساعات التوافر ÷ إجمالي ساعات التشغيل × 100",
        "unit": "%",
        "frequency": "شهري",
        "owner": "قسم تقنية المعلومات",
        "data_owner": "قسم تقنية المعلومات",
        "approver": "وحدة الاستراتيجية والتميز المؤسسي",
        "source": "سجلات النظام / الاستضافة",
        "target_2026": "99%",
        "priority": "عال",
        "recommendation": "يضاف للنظام",
    },
    {
        "program_code": "PRG-2026-007",
        "program": "برنامج التحول الرقمي ودعم نظام الجودة",
        "indicator_code": "QOP-2026-014",
        "indicator": "متوسط زمن الاستجابة للدعم الفني",
        "definition": "متوسط الزمن من فتح طلب الدعم الفني إلى أول استجابة موثقة.",
        "formula": "مجموع أزمنة الاستجابة ÷ عدد طلبات الدعم",
        "unit": "ساعة",
        "frequency": "شهري",
        "owner": "قسم تقنية المعلومات",
        "data_owner": "قسم تقنية المعلومات",
        "approver": "إدارة الدعم المؤسسي",
        "source": "سجل طلبات الدعم الفني",
        "target_2026": "≤ 8 ساعات عمل",
        "priority": "متوسط",
        "recommendation": "مرحلة ثانية",
    },
]


HEADERS = [
    ("program_code", "كود البرنامج"),
    ("program", "البرنامج"),
    ("indicator_code", "كود المؤشر المقترح"),
    ("indicator", "المؤشر التشغيلي"),
    ("definition", "التعريف"),
    ("formula", "طريقة القياس"),
    ("unit", "الوحدة"),
    ("frequency", "الدورية"),
    ("owner", "مالك الأداء"),
    ("data_owner", "مالك البيانات"),
    ("approver", "جهة الاعتماد"),
    ("source", "مصدر البيانات"),
    ("target_2026", "مستهدف 2026"),
    ("priority", "الأولوية"),
    ("recommendation", "التوصية"),
]


def write_markdown(path: Path) -> None:
    lines = [
        "# كتالوج المؤشرات التشغيلية الداخلية لبرامج نظام الجودة",
        "",
        f"**التاريخ:** {DATE}",
        "",
        "## قاعدة التصنيف",
        "",
        "- هذه المؤشرات تخص برامج نظام الجودة فقط.",
        "- لا تنسخ بيانات برامج المستفيدين الموجودة في رافد.",
        "- المؤشرات ذات توصية `يضاف للنظام` هي الدفعة الأولى المقترحة.",
        "- المؤشرات ذات توصية `مرحلة ثانية` تؤجل حتى تتضح مصادر القياس أو آلية جمع البيانات.",
        "",
        "## المؤشرات المقترحة",
        "",
        "| البرنامج | كود المؤشر | المؤشر | الدورية | المالك | مصدر البيانات | مستهدف 2026 | التوصية |",
        "|---|---|---|---|---|---|---|---|",
    ]
    for row in ROWS:
        lines.append(
            "| "
            + " | ".join(
                [
                    row["program"],
                    row["indicator_code"],
                    row["indicator"],
                    row["frequency"],
                    row["owner"],
                    row["source"],
                    row["target_2026"],
                    row["recommendation"],
                ]
            )
            + " |"
        )

    first_wave = [row for row in ROWS if row["recommendation"] == "يضاف للنظام"]
    second_wave = [row for row in ROWS if row["recommendation"] != "يضاف للنظام"]

    lines.extend(
        [
            "",
            "## الدفعة الأولى المقترحة للإضافة",
            "",
            *[f"- `{row['indicator_code']}`: {row['indicator']}" for row in first_wave],
            "",
            "## مؤشرات المرحلة الثانية",
            "",
            *[f"- `{row['indicator_code']}`: {row['indicator']}" for row in second_wave],
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def write_json(path: Path) -> None:
    payload = {
        "title": "كتالوج المؤشرات التشغيلية الداخلية لبرامج نظام الجودة",
        "date": DATE,
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "policy": "مؤشرات نظام الجودة فقط؛ لا تكرار لبيانات رافد التشغيلية الخاصة بالمستفيدين.",
        "rows": ROWS,
        "firstWaveCount": sum(1 for row in ROWS if row["recommendation"] == "يضاف للنظام"),
        "secondWaveCount": sum(1 for row in ROWS if row["recommendation"] != "يضاف للنظام"),
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def style_sheet(ws) -> None:
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
            cell.border = border
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font


def autosize(ws, widths: dict[int, int] | None = None) -> None:
    widths = widths or {}
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        if col in widths:
            ws.column_dimensions[letter].width = widths[col]
            continue
        width = 12
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row, col).value
            if value:
                width = max(width, min(55, len(str(value)) + 3))
        ws.column_dimensions[letter].width = width


def write_excel(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "كتالوج المؤشرات"
    ws.append([label for _, label in HEADERS])
    for row in ROWS:
        ws.append([row[key] for key, _ in HEADERS])
    style_sheet(ws)
    autosize(
        ws,
        {
            1: 18,
            2: 34,
            3: 20,
            4: 34,
            5: 48,
            6: 42,
            7: 12,
            8: 16,
            9: 28,
            10: 28,
            11: 24,
            12: 32,
            13: 22,
            14: 14,
            15: 18,
        },
    )

    ws2 = wb.create_sheet("دفعة أولى")
    ws2.append([label for _, label in HEADERS])
    for row in ROWS:
        if row["recommendation"] == "يضاف للنظام":
            ws2.append([row[key] for key, _ in HEADERS])
    style_sheet(ws2)
    autosize(ws2)

    ws3 = wb.create_sheet("مرحلة ثانية")
    ws3.append([label for _, label in HEADERS])
    for row in ROWS:
        if row["recommendation"] != "يضاف للنظام":
            ws3.append([row[key] for key, _ in HEADERS])
    style_sheet(ws3)
    autosize(ws3)

    wb.save(path)


def main() -> None:
    DOCS.mkdir(parents=True, exist_ok=True)
    PLANS.mkdir(parents=True, exist_ok=True)
    md_path = DOCS / f"qms-operational-indicators-catalog-{DATE}.md"
    json_path = DOCS / f"qms-operational-indicators-catalog-{DATE}.json"
    xlsx_path = PLANS / f"كتالوج_المؤشرات_التشغيلية_لبرامج_نظام_الجودة_{DATE}.xlsx"

    write_markdown(md_path)
    write_json(json_path)
    write_excel(xlsx_path)

    print(f"Markdown: {md_path}")
    print(f"JSON: {json_path}")
    print(f"Excel: {xlsx_path}")
    print(f"Indicators: {len(ROWS)}")
    print(f"First wave: {sum(1 for row in ROWS if row['recommendation'] == 'يضاف للنظام')}")


if __name__ == "__main__":
    main()
