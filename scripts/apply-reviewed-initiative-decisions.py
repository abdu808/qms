import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
TODAY = "2026-05-02"
PLAN_DIR = ROOT / "ISO9001" / "الخطط والمشرات"
SOURCE = PLAN_DIR / "المبادرات_الـ21_2026_2030.xlsx"
OUTPUT = PLAN_DIR / f"المبادرات_المعتمدة_بعد_المراجعة_2026_2030_{TODAY}.xlsx"
DECISIONS_MD = ROOT / "docs" / f"initiative-review-decisions-{TODAY}.md"
DECISIONS_JSON = ROOT / "docs" / f"initiative-review-decisions-{TODAY}.json"


DELETE_DECISIONS = {
    "INI-2026-015": "حذف: حملة الزكاة المؤسسية غير مناسبة لأن الزكاة الإجبارية تذهب إلى هيئة الزكاة.",
    "INI-2026-025": "حذف من الملف: شراكة بنك التنمية لا تنشأ ما لم يصدر قرار جديد.",
    "INI-2026-026": "حذف: لا حاجة لشراكة TVTC بوجود معهد تدريب تابع للجمعية.",
    "INI-2026-030": "حذف: تأسيس وظيفة إدارة علاقات الشركاء ملفه وتعقيده لا يستحقان في هذه المرحلة.",
}

UPDATES = {
    "INI-2026-004": {
        "اسم المبادرة": "4 حملات تبرع كبرى موسمية محكمة التخطيط والتنفيذ",
        "المالك": "فاطمة عقيبي",
        "ملاحظات المراجعة": "اعتمدت صياغة النظام بدلاً من 6 حملات؛ الأفضل جودة التنفيذ لا كثرة الحملات.",
    },
    "INI-2026-003": {
        "اسم المبادرة": "تخصيص بند سنوي لبناء محفظة استثمارية آمنة من الإيرادات غير المقيدة وعوائد الاستثمارات",
        "الحالة": "ON_HOLD",
        "ملاحظات المراجعة": "التمويل يكون فقط من الإيرادات غير المقيدة وعوائد الاستثمارات، ولا يشمل الزكاة أو التبرعات المقيدة أو مخصصات المستفيدين. التحويل للمحفظة يتم بعد مراجعة السيولة والالتزامات واعتماد مجلس الإدارة.",
    },
    "INI-2026-005": {
        "اسم المبادرة": "كفالة 850 يتيم مكفول مع زيادة سنوية ثابتة وتحسين أثر الكفالة",
        "المالك": "خاتمة محرق",
        "ملاحظات المراجعة": "المالك الصحيح خاتمة. يمكن لاحقاً فتح نقاش مستقل لمؤشرين: الإبقاء على الكافلين، وتحسين مصروف/أثر الكفالة.",
    },
    "INI-2026-006": {
        "المالك": "طلال الحربي",
        "ملاحظات المراجعة": "تُنقل لاختصاص إدارة المساعدات العينية.",
    },
    "INI-2026-010": {
        "اسم المبادرة": "إطلاق نموذج قياس أثر يُطبَّق على 50 أسرة في 2026 تدرجاً نحو 200 أسرة في السنوات اللاحقة",
        "المالك": "ايلاف حسن",
        "ملاحظات المراجعة": "اعتمدت صياغة النظام لأنها أكثر واقعية وقابلة للتدرج.",
    },
    "INI-2026-007": {
        "المالك": "خاتمة محرق",
        "ملاحظات المراجعة": "معتمدة مع اعتماد مالك النظام.",
    },
    "INI-2026-008": {
        "المالك": "نادية قلم",
        "ملاحظات المراجعة": "معتمدة مع اعتماد مالك النظام.",
    },
    "INI-2026-009": {
        "المالك": "خاتمة محرق",
        "ملاحظات المراجعة": "معتمدة مع اعتماد مالك النظام.",
    },
    "INI-2026-012": {
        "المالك": "نادية قلم",
        "ملاحظات المراجعة": "معتمدة مع اعتماد مالك النظام.",
    },
    "INI-2026-017": {
        "اسم المبادرة": "تعزيز النضج الرقمي وتكامل الأنظمة التشغيلية ودعم القرار المؤسسي",
        "المالك": "عبدالرحمن عقيل",
        "ملاحظات المراجعة": "أعيدت الصياغة لتصبح مبادرة أوسع من مجرد تشغيل Rafid ERP وتكامل QMS.",
    },
    "INI-2026-018": {
        "المالك": "خليل هادي",
        "ملاحظات المراجعة": "معتمدة مع اعتماد مالك النظام.",
    },
    "INI-2026-023": {
        "اسم المبادرة": "مراجعة وتطوير الشراكات القائمة ورفع فاعليتها وفق معايير واضحة",
        "ملاحظات المراجعة": "صياغة ألطف من تدقيق الشراكات وفق المعايير الصارمة.",
    },
    "INI-2026-031": {
        "اسم المبادرة": "24 مادة إعلامية سنوياً مرتبطة بالمشاريع والأثر مع نمو المتابعين 20%",
        "ملاحظات المراجعة": "اعتمدت صياغة النظام بدلاً من 50 مادة و30% نمو لأنها أكثر قابلية للتنفيذ.",
    },
}

ADDITIONS = [
    {
        "الكود": "INI-2026-032",
        "اسم المبادرة": "تشغيل منظومة قياس تجربة المستفيد عبر QMS: استبيانات رضا ربعية، وتتبع SLA، ولوحة متابعة فورية",
        "الهدف الاستراتيجي": "STR-2026-006",
        "الحالة": "NOT_STARTED",
        "التقدم %": "0%",
        "الميزانية": "—",
        "المُنفَق": "—",
        "تاريخ البداية": "2026-01-01",
        "المالك": "خاتمة محرق",
        "ملاحظات المراجعة": "معتمدة ومضافة للملف لأنها موجودة في النظام وتخدم تجربة المستفيد.",
    },
    {
        "الكود": "INI-2026-033",
        "اسم المبادرة": "مسار الحصول على شهادة ISO 9001:2015 خلال 2026 عبر Stage 1 وStage 2 مع جهة معتمدة",
        "الهدف الاستراتيجي": "STR-2026-012",
        "الحالة": "IN_PROGRESS",
        "التقدم %": "0%",
        "الميزانية": "—",
        "المُنفَق": "—",
        "تاريخ البداية": "2026-01-01",
        "المالك": "ايلاف حسن",
        "ملاحظات المراجعة": "معتمدة ومضافة للملف لأنها مبادرة مباشرة لشهادة ISO.",
    },
]


def read_source_rows():
    wb = load_workbook(SOURCE, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = 3
    headers = [cell.value for cell in ws[header_row] if cell.value is not None]
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        if item.get("الكود"):
            rows.append(item)
    return headers, rows


def style_header(ws):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autofit(ws):
    for col_idx, col in enumerate(ws.columns, start=1):
        width = 10
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(55, len(value) + 3))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    style_header(ws)
    for idx, row in enumerate(rows, start=1):
        row = dict(row)
        row["#"] = idx if "#" in headers else row.get("#")
        ws.append([row.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    autofit(ws)
    return ws


def main():
    headers, source_rows = read_source_rows()
    if "ملاحظات المراجعة" not in headers:
        headers = headers + ["ملاحظات المراجعة"]

    active = []
    archived = []
    decisions = []

    for row in source_rows:
        code = row.get("الكود")
        row = dict(row)
        if code in DELETE_DECISIONS:
            row["ملاحظات المراجعة"] = DELETE_DECISIONS[code]
            archived.append(row)
            decisions.append({"code": code, "decision": "DELETE_FROM_APPROVED_FILE", "note": DELETE_DECISIONS[code]})
            continue
        if code in UPDATES:
            row.update(UPDATES[code])
            decisions.append({"code": code, "decision": "UPDATE", "updates": UPDATES[code]})
        else:
            row["ملاحظات المراجعة"] = row.get("ملاحظات المراجعة") or "معتمدة كما هي"
            decisions.append({"code": code, "decision": "APPROVE"})
        active.append(row)

    for row in ADDITIONS:
        active.append(dict(row))
        decisions.append({"code": row["الكود"], "decision": "ADD_TO_APPROVED_FILE", "row": row})

    active.sort(key=lambda r: str(r.get("الكود")))
    archived.sort(key=lambda r: str(r.get("الكود")))

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "المبادرات المعتمدة", headers, active)
    write_sheet(wb, "المبادرات المحذوفة", headers, archived)
    write_sheet(
        wb,
        "ملخص القرارات",
        ["البند", "القيمة"],
        [
            {"البند": "تاريخ التحديث", "القيمة": datetime.now().strftime("%Y-%m-%d %H:%M")},
            {"البند": "المبادرات المعتمدة", "القيمة": len(active)},
            {"البند": "المبادرات المحذوفة/المؤرشفة", "القيمة": len(archived)},
            {"البند": "المبادرات المضافة", "القيمة": len(ADDITIONS)},
            {"البند": "الملف المصدر", "القيمة": str(SOURCE.relative_to(ROOT)).replace("\\", "/")},
        ],
    )
    wb.save(OUTPUT)

    payload = {
        "output": str(OUTPUT.relative_to(ROOT)).replace("\\", "/"),
        "approvedCount": len(active),
        "archivedCount": len(archived),
        "addedCount": len(ADDITIONS),
        "decisions": decisions,
        "approvedCodes": [r["الكود"] for r in active],
        "archivedCodes": [r["الكود"] for r in archived],
    }
    DECISIONS_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    md = f"""# قرارات مراجعة المبادرات

**التاريخ:** {datetime.now().strftime("%Y-%m-%d %H:%M")}

## النتيجة

- المبادرات المعتمدة في النسخة الجديدة: {len(active)}
- المبادرات المحذوفة/المؤرشفة من الملف المعتمد: {len(archived)}
- المبادرات المضافة للملف: {len(ADDITIONS)}
- الملف الناتج: `{str(OUTPUT.relative_to(ROOT)).replace("\\", "/")}`

## المبادرات المحذوفة/المؤرشفة

{chr(10).join(f'- `{r["الكود"]}`: {r.get("اسم المبادرة")} — {r.get("ملاحظات المراجعة")}' for r in archived)}

## ملاحظات تحتاج نقاش لاحق

- `INI-2026-005`: يمكن لاحقاً فصلها إلى مؤشرين أو مبادرتين فرعيتين: الإبقاء على الكافلين، وتحسين مصروف/أثر الكفالة.
- حذف/أرشفة المبادرات من النظام الحي، إن رغبت به لاحقاً، يحتاج تأكيداً مستقلاً قبل التنفيذ لأنه يغيّر بيانات الإنتاج.
"""
    DECISIONS_MD.write_text(md, encoding="utf-8")

    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
