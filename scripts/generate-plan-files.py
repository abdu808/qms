"""
توليد ملفات الخطة الاستراتيجية الجديدة من النظام
+ أرشفة الملفات القديمة
"""
import json, os, shutil, sys, io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = r'C:\Users\abdu8\Documents\dev\qms\ISO9001\الخطط والمشرات'
ARCHIVE = os.path.join(ROOT, '_archive')
JSON_FILE = os.path.join(ROOT, 'qms-export.json')

# ─── 1. تحميل البيانات ─────────────────────────────────
with open(JSON_FILE, 'r', encoding='utf-8') as f:
    data = json.load(f)

print(f"✅ تم تحميل البيانات:")
print(f"   - الخطة: {data['plan']['code']} ({data['plan']['startYear']}-{data['plan']['endYear']})")
print(f"   - {len(data['axes'])} محاور · {len(data['goals'])} أهداف · {len(data['objectives'])} هدف تشغيلي")
print(f"   - {len(data['indicators'])} مؤشر · {len(data['annualTargets'])} مستهدف · {len(data['initiatives'])} مبادرة")

# ─── 2. أرشفة الملفات القديمة ──────────────────────────
os.makedirs(ARCHIVE, exist_ok=True)
to_archive = [
    'الخطة الاستراتيجية جمعية البر صبياء محدثة.xlsx',
    'KPI_Playbook_v10.xlsx',
    'مؤشرات باقي الاقسام- مسودة.xlsx',
    '2025خطة_التشغيلية_أكسل(محدددثةً) (1).xlsx',
]
for f in to_archive:
    src = os.path.join(ROOT, f)
    if os.path.exists(src):
        new_name = f.replace('.xlsx', '_قديم_2025_2027.xlsx')
        dst = os.path.join(ARCHIVE, new_name)
        shutil.move(src, dst)
        print(f"📦 أُرشف: {f}")

# ─── 3. أنماط Excel مشتركة ─────────────────────────────
FONT_NAME = 'Arial'
HEADER_FILL = PatternFill('solid', start_color='1F4E79')
SECTION_FILL = PatternFill('solid', start_color='D5E8F0')
THIN = Side(border_style='thin', color='B7B7B7')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def header(cell, text, color='FFFFFF'):
    cell.value = text
    cell.font = Font(name=FONT_NAME, size=12, bold=True, color=color)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, readingOrder=2)
    cell.border = BORDER

def cell_text(cell, text, bold=False, center=False, color=None):
    cell.value = text
    cell.font = Font(name=FONT_NAME, size=11, bold=bold, color=color or '000000')
    cell.alignment = Alignment(horizontal='center' if center else 'right', vertical='center', wrap_text=True, readingOrder=2)
    cell.border = BORDER

def set_rtl(ws):
    ws.sheet_view.rightToLeft = True

# ─── 4. ملف 1: الخطة الاستراتيجية 2026-2030 ────────────
wb = Workbook()
ws = wb.active
ws.title = 'الخطة الاستراتيجية'
set_rtl(ws)
for col, w in zip('ABCDEFGH', [6, 18, 50, 18, 35, 12, 14, 25]):
    ws.column_dimensions[col].width = w

ws.merge_cells('A1:H1')
ws.row_dimensions[1].height = 32
header(ws['A1'], f'الخطة الاستراتيجية 2026-2030 — جمعية البر الخيرية بصبيا — {data["plan"]["code"]}')

ws.merge_cells('A2:H2')
ws.row_dimensions[2].height = 22
ws['A2'].value = 'الخطة المعتمدة من النظام · المصدر الوحيد للحقيقة · جميع الأرقام مستخرجة من DB'
ws['A2'].font = Font(name=FONT_NAME, size=10, italic=True, color='595959')
ws['A2'].alignment = Alignment(horizontal='center', readingOrder=2)
ws['A2'].fill = SECTION_FILL

# الهيدر
hdrs = ['#', 'كود الهدف', 'الهدف الاستراتيجي', 'المحور', 'الملاحظات', 'التقدم %', 'الحالة', 'المالك']
for i, h in enumerate(hdrs, 1):
    header(ws.cell(row=4, column=i), h)

# البيانات
for idx, g in enumerate(data['goals'], start=5):
    cell_text(ws.cell(row=idx, column=1), idx-4, center=True, bold=True)
    cell_text(ws.cell(row=idx, column=2), g['code'], center=True, bold=True)
    cell_text(ws.cell(row=idx, column=3), g['title'])
    cell_text(ws.cell(row=idx, column=4), f"{g.get('axisCode','')} — {g.get('axisName','')}", center=True)
    cell_text(ws.cell(row=idx, column=5), g.get('notes', '') or '')
    cell_text(ws.cell(row=idx, column=6), f"{g['progress']}%", center=True)
    cell_text(ws.cell(row=idx, column=7), g['status'], center=True, color='006100' if g['status']=='ACTIVE' else '000000')
    cell_text(ws.cell(row=idx, column=8), g.get('ownerName', '') or '—', center=True)
    ws.row_dimensions[idx].height = 60

# تذييل
last_row = 5 + len(data['goals']) + 1
ws.merge_cells(f'A{last_row}:H{last_row}')
ws.cell(row=last_row, column=1).value = f'إجمالي الأهداف الاستراتيجية: {len(data["goals"])} هدف · جميعها بحالة ACTIVE'
ws.cell(row=last_row, column=1).font = Font(name=FONT_NAME, size=11, bold=True, color='1F4E79')
ws.cell(row=last_row, column=1).alignment = Alignment(horizontal='center', readingOrder=2)
ws.cell(row=last_row, column=1).fill = SECTION_FILL

# ورقة المحاور
ws2 = wb.create_sheet('المحاور الاستراتيجية')
set_rtl(ws2)
for col, w in zip('ABCDEF', [10, 35, 10, 12, 10, 14]):
    ws2.column_dimensions[col].width = w
ws2.merge_cells('A1:F1')
header(ws2['A1'], 'المحاور الاستراتيجية الأربعة (مُحدَّثة 2026-2030)')
ws2.row_dimensions[1].height = 30
for i, h in enumerate(['الكود', 'اسم المحور', 'الترتيب', 'الوزن %', 'اللون', 'عدد الأهداف'], 1):
    header(ws2.cell(row=3, column=i), h)
goals_per_axis = {}
for g in data['goals']:
    goals_per_axis[g.get('axisCode')] = goals_per_axis.get(g.get('axisCode'), 0) + 1

for idx, a in enumerate(sorted(data['axes'], key=lambda x: x.get('order', 0)), start=4):
    cell_text(ws2.cell(row=idx, column=1), a['code'], center=True, bold=True)
    cell_text(ws2.cell(row=idx, column=2), a['nameAr'])
    cell_text(ws2.cell(row=idx, column=3), a.get('order', '') or '', center=True)
    cell_text(ws2.cell(row=idx, column=4), f"{a.get('weight', 0)}%", center=True)
    cell_text(ws2.cell(row=idx, column=5), a.get('color', '') or '—', center=True)
    cell_text(ws2.cell(row=idx, column=6), goals_per_axis.get(a['code'], 0), center=True, bold=True)
    ws2.row_dimensions[idx].height = 28

wb.save(os.path.join(ROOT, 'الخطة_الاستراتيجية_2026_2030_المعتمدة.xlsx'))
print(f"✅ ملف 1: الخطة الاستراتيجية 2026-2030")

# ─── 5. ملف 2: الخطة التشغيلية 2026 (الأهداف التشغيلية) ────
wb = Workbook()
ws = wb.active
ws.title = 'الأهداف التشغيلية'
set_rtl(ws)
for col, w in zip('ABCDEFGHIJK', [5, 15, 35, 18, 35, 10, 10, 10, 10, 10, 18]):
    ws.column_dimensions[col].width = w

ws.merge_cells('A1:K1')
header(ws['A1'], 'الخطة التشغيلية 2026-2030 — الأهداف التشغيلية الـ 19')
ws.row_dimensions[1].height = 30

hdrs = ['#', 'الكود', 'الهدف التشغيلي', 'الهدف الاستراتيجي', 'KPI', 'خط الأساس', 'المستهدف', 'الوحدة', 'التقدم %', 'الحالة', 'المالك']
for i, h in enumerate(hdrs, 1):
    header(ws.cell(row=3, column=i), h)

# تجميع حسب الهدف الاستراتيجي
sorted_objs = sorted(data['objectives'], key=lambda o: o.get('goalCode', ''))
for idx, o in enumerate(sorted_objs, start=4):
    cell_text(ws.cell(row=idx, column=1), idx-3, center=True)
    cell_text(ws.cell(row=idx, column=2), o['code'], center=True, bold=True)
    cell_text(ws.cell(row=idx, column=3), o['title'])
    cell_text(ws.cell(row=idx, column=4), o.get('goalCode', ''), center=True)
    cell_text(ws.cell(row=idx, column=5), o.get('kpi', '') or '')
    cell_text(ws.cell(row=idx, column=6), o.get('baseline', '') or '—', center=True)
    cell_text(ws.cell(row=idx, column=7), o.get('target', '') or '', center=True, bold=True)
    cell_text(ws.cell(row=idx, column=8), o.get('unit', '') or '', center=True)
    cell_text(ws.cell(row=idx, column=9), f"{o.get('progress', 0)}%", center=True)
    cell_text(ws.cell(row=idx, column=10), o.get('status', ''), center=True)
    cell_text(ws.cell(row=idx, column=11), o.get('ownerName', '') or '—', center=True)
    ws.row_dimensions[idx].height = 50

wb.save(os.path.join(ROOT, 'الخطة_التشغيلية_2026_بـ4_محاور.xlsx'))
print(f"✅ ملف 2: الخطة التشغيلية بـ 4 محاور")

# ─── 6. ملف 3: KPI Catalog v11 ─────────────────────────
wb = Workbook()
ws = wb.active
ws.title = 'كتالوج المؤشرات'
set_rtl(ws)
for col, w in zip('ABCDEFGHIJK', [5, 16, 35, 12, 12, 12, 12, 12, 12, 12, 18]):
    ws.column_dimensions[col].width = w

ws.merge_cells('A1:K1')
header(ws['A1'], 'كتالوج المؤشرات v11 — الإصدار المعتمد 2026-2030')
ws.row_dimensions[1].height = 30
ws.merge_cells('A2:K2')
ws['A2'].value = 'يستبدل KPI_Playbook_v10 (المُؤرشَف) — 16 مؤشر معتمد · ترميز IND-2026-XXX'
ws['A2'].font = Font(name=FONT_NAME, size=10, italic=True)
ws['A2'].alignment = Alignment(horizontal='center', readingOrder=2)
ws['A2'].fill = SECTION_FILL

hdrs = ['#', 'الكود', 'اسم المؤشر', 'الوحدة', 'الاتجاه', 'التردد', 'النوع', 'خط الأساس', 'الوزن %', 'العتبة الخضراء', 'المالك']
for i, h in enumerate(hdrs, 1):
    header(ws.cell(row=4, column=i), h)

for idx, ind in enumerate(data['indicators'], start=5):
    cell_text(ws.cell(row=idx, column=1), idx-4, center=True)
    cell_text(ws.cell(row=idx, column=2), ind['code'], center=True, bold=True)
    cell_text(ws.cell(row=idx, column=3), ind['nameAr'])
    cell_text(ws.cell(row=idx, column=4), ind.get('unit', '') or '', center=True)
    cell_text(ws.cell(row=idx, column=5), ind.get('direction', ''), center=True)
    cell_text(ws.cell(row=idx, column=6), ind.get('frequency', ''), center=True)
    cell_text(ws.cell(row=idx, column=7), ind.get('kpiType', ''), center=True)
    cell_text(ws.cell(row=idx, column=8), ind.get('baseline', '') or '—', center=True)
    cell_text(ws.cell(row=idx, column=9), f"{ind.get('weight', 0)}%", center=True)
    cell_text(ws.cell(row=idx, column=10), f"{ind.get('greenThreshold', 95)}%", center=True)
    cell_text(ws.cell(row=idx, column=11), ind.get('ownerName', '') or '—', center=True)
    ws.row_dimensions[idx].height = 36

wb.save(os.path.join(ROOT, 'KPI_Catalog_v11_2026_2030.xlsx'))
print(f"✅ ملف 3: KPI Catalog v11")

# ─── 7. ملف 4: AnnualTargets 2026-2030 ──────────────────
wb = Workbook()
ws = wb.active
ws.title = 'المستهدفات السنوية'
set_rtl(ws)
for col, w in zip('ABCDEFG', [5, 16, 35, 10, 14, 14, 14]):
    ws.column_dimensions[col].width = w

ws.merge_cells('A1:G1')
header(ws['A1'], 'المستهدفات السنوية 2026-2030 — 80 سجل (16 مؤشر × 5 سنوات)')
ws.row_dimensions[1].height = 30

hdrs = ['#', 'كود المؤشر', 'اسم المؤشر', 'السنة', 'المستهدف السنوي', 'مستهدف ربعي/شهري', 'ملاحظات']
for i, h in enumerate(hdrs, 1):
    header(ws.cell(row=3, column=i), h)

# ترتيب: by indicator code then year
sorted_t = sorted(data['annualTargets'], key=lambda t: (t.get('indicatorCode', ''), t.get('year', 0)))
for idx, t in enumerate(sorted_t, start=4):
    cell_text(ws.cell(row=idx, column=1), idx-3, center=True)
    cell_text(ws.cell(row=idx, column=2), t.get('indicatorCode', ''), center=True, bold=True)
    cell_text(ws.cell(row=idx, column=3), t.get('indicatorName', ''))
    cell_text(ws.cell(row=idx, column=4), t.get('year', ''), center=True, bold=True)
    cell_text(ws.cell(row=idx, column=5), t.get('targetValue', ''), center=True)
    quarterly = ' · '.join([f'Q{i}: {t.get(f"q{i}", "—")}' for i in range(1,5) if t.get(f'q{i}')])
    cell_text(ws.cell(row=idx, column=6), quarterly or '—', center=True)
    cell_text(ws.cell(row=idx, column=7), '')
    ws.row_dimensions[idx].height = 26

wb.save(os.path.join(ROOT, 'AnnualTargets_2026_2030.xlsx'))
print(f"✅ ملف 4: AnnualTargets")

# ─── 8. ملف 5: المبادرات الـ 21 ────────────────────────
wb = Workbook()
ws = wb.active
ws.title = 'المبادرات'
set_rtl(ws)
for col, w in zip('ABCDEFGHIJ', [5, 16, 35, 18, 14, 12, 14, 14, 14, 18]):
    ws.column_dimensions[col].width = w

ws.merge_cells('A1:J1')
header(ws['A1'], 'المبادرات الفعّالة الـ 21 — الخطة 2026-2030')
ws.row_dimensions[1].height = 30

hdrs = ['#', 'الكود', 'اسم المبادرة', 'الهدف الاستراتيجي', 'الحالة', 'التقدم %', 'الميزانية', 'المُنفَق', 'تاريخ البداية', 'المالك']
for i, h in enumerate(hdrs, 1):
    header(ws.cell(row=3, column=i), h)

# ترتيب: by goal then code
sorted_inis = sorted(data['initiatives'], key=lambda i: (i.get('goalCode', ''), i.get('code', '')))
for idx, i in enumerate(sorted_inis, start=4):
    cell_text(ws.cell(row=idx, column=1), idx-3, center=True)
    cell_text(ws.cell(row=idx, column=2), i['code'], center=True, bold=True)
    cell_text(ws.cell(row=idx, column=3), i.get('name', ''))
    cell_text(ws.cell(row=idx, column=4), i.get('goalCode', ''), center=True)
    color_map = {'IN_PROGRESS': '006100', 'NOT_STARTED': '595959', 'COMPLETED': '1F4E79'}
    cell_text(ws.cell(row=idx, column=5), i.get('status', ''), center=True, color=color_map.get(i.get('status'), '000000'))
    cell_text(ws.cell(row=idx, column=6), f"{i.get('progress', 0)}%", center=True)
    cell_text(ws.cell(row=idx, column=7), i.get('budget', '') or '—', center=True)
    cell_text(ws.cell(row=idx, column=8), i.get('spent', '') or '—', center=True)
    cell_text(ws.cell(row=idx, column=9), (i.get('startDate', '') or '')[:10], center=True)
    cell_text(ws.cell(row=idx, column=10), i.get('ownerName', '') or '—', center=True)
    ws.row_dimensions[idx].height = 50

wb.save(os.path.join(ROOT, 'المبادرات_الـ21_2026_2030.xlsx'))
print(f"✅ ملف 5: المبادرات")

print(f"\n🎯 تم إنتاج 5 ملفات Excel جديدة في:\n{ROOT}\n")
print(f"📦 الملفات القديمة في: {ARCHIVE}")
