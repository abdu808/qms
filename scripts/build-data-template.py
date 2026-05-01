"""
بناء ملف Excel لتعبئة بيانات الخطة الاستراتيجية الفعلية
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ────── أنماط مشتركة ──────
FONT = 'Arial'
HEADER_FILL = PatternFill('solid', start_color='1F4E79')
SECTION_FILL = PatternFill('solid', start_color='D5E8F0')
INPUT_FILL = PatternFill('solid', start_color='FFF2CC')   # أصفر للحقول المطلوبة
NOTES_FILL = PatternFill('solid', start_color='E7E6E6')   # رمادي للملاحظات
CORRECT_FILL = PatternFill('solid', start_color='E2EFDA')  # أخضر للصحيح

THIN = Side(border_style='thin', color='B7B7B7')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def header(cell, text, color='FFFFFF'):
    cell.value = text
    cell.font = Font(name=FONT, size=12, bold=True, color=color)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)
    cell.border = BORDER

def section(cell, text):
    cell.value = text
    cell.font = Font(name=FONT, size=13, bold=True, color='1F4E79')
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)
    cell.border = BORDER

def label(cell, text):
    cell.value = text
    cell.font = Font(name=FONT, size=11)
    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)
    cell.border = BORDER

def input_cell(cell, hint=''):
    cell.value = hint
    cell.font = Font(name=FONT, size=11, color='808080', italic=True)
    cell.fill = INPUT_FILL
    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)
    cell.border = BORDER

def note(cell, text):
    cell.value = text
    cell.font = Font(name=FONT, size=10, italic=True, color='595959')
    cell.fill = NOTES_FILL
    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)
    cell.border = BORDER

def correct(cell, text):
    cell.value = text
    cell.font = Font(name=FONT, size=11, color='006100')
    cell.fill = CORRECT_FILL
    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)
    cell.border = BORDER

def set_rtl(ws):
    ws.sheet_view.rightToLeft = True

# ════════════════════════════════════════════════════════
# الورقة 1: التعليمات
# ════════════════════════════════════════════════════════
ws = wb.active
ws.title = '0_التعليمات'
set_rtl(ws)

ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 80
ws.row_dimensions[1].height = 30

ws.merge_cells('A1:B1')
header(ws['A1'], 'ملف بيانات الخطة الاستراتيجية 2026-2030 — جمعية البر الخيرية بصبيا')

ws['B3'] = 'التاريخ: 2026-04-30'
ws['B3'].font = Font(name=FONT, size=11)
ws['B3'].alignment = Alignment(horizontal='right', readingOrder=2)

instructions = [
    ('', ''),
    ('A1', 'كيفية تعبئة الملف'),
    ('S', 'هذا الملف يحتوي على 6 أوراق عمل لتعبئة البيانات الفعلية للخطة الاستراتيجية.'),
    ('S', 'الخلايا الصفراء 🟡 هي حقول التعبئة المطلوبة منكم.'),
    ('S', 'الخلايا الرمادية ملاحظات توضيحية — لا تُعدَّل.'),
    ('S', 'الخلايا الخضراء قيم محسوبة أو ثابتة معتمدة.'),
    ('', ''),
    ('A2', 'محتوى الأوراق'),
    ('S', '1_المستفيدين: قاعدة المستفيدين الفعلية 2025'),
    ('S', '2_الكفالة: برنامج كفالة الأيتام'),
    ('S', '3_الأرقام_المالية: إيرادات ومصروفات 2025'),
    ('S', '4_الشراكات: قائمة الشركاء الـ 18'),
    ('S', '5_الكوادر: الموظفون والمتطوعون'),
    ('S', '6_المالكون_والمبادرات: تصحيحات الملكية'),
    ('', ''),
    ('A3', 'بعد التعبئة'),
    ('S', '1. احفظ الملف بنفس الاسم'),
    ('S', '2. أرسله للمستشار للتحديث في النظام'),
    ('S', '3. سيُنفّذ التحديث في DB + وثيقة المجلس النهائية'),
    ('', ''),
    ('A4', 'ترميز الألوان'),
    ('Y', 'حقل تعبئة مطلوب — أدخل القيمة هنا'),
    ('G', 'قيمة معتمدة أو محسوبة — لا تُعدَّل'),
    ('N', 'ملاحظة توضيحية — مرجع فقط'),
]

row = 5
for typ, text in instructions:
    if typ == '':
        row += 1
        continue
    ws.row_dimensions[row].height = 22
    if typ == 'A1':
        ws.merge_cells(f'A{row}:B{row}')
        section(ws[f'A{row}'], text)
    elif typ == 'A2' or typ == 'A3' or typ == 'A4':
        ws.merge_cells(f'A{row}:B{row}')
        section(ws[f'A{row}'], text)
    elif typ == 'S':
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'].value = text
        ws[f'A{row}'].font = Font(name=FONT, size=11)
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)
    elif typ == 'Y':
        input_cell(ws[f'A{row}'], '')
        ws[f'B{row}'].value = text
        ws[f'B{row}'].font = Font(name=FONT, size=11)
        ws[f'B{row}'].alignment = Alignment(horizontal='right', readingOrder=2)
    elif typ == 'G':
        correct(ws[f'A{row}'], '')
        ws[f'B{row}'].value = text
        ws[f'B{row}'].font = Font(name=FONT, size=11)
        ws[f'B{row}'].alignment = Alignment(horizontal='right', readingOrder=2)
    elif typ == 'N':
        note(ws[f'A{row}'], '')
        ws[f'B{row}'].value = text
        ws[f'B{row}'].font = Font(name=FONT, size=11)
        ws[f'B{row}'].alignment = Alignment(horizontal='right', readingOrder=2)
    row += 1

# ════════════════════════════════════════════════════════
# الورقة 2: المستفيدين
# ════════════════════════════════════════════════════════
ws = wb.create_sheet('1_المستفيدين')
set_rtl(ws)

ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 50

ws.merge_cells('A1:C1')
header(ws['A1'], 'قاعدة المستفيدين الفعلية 2025')

label(ws['A3'], 'البند')
label(ws['B3'], 'القيمة')
label(ws['C3'], 'ملاحظات / منهجية')
for cell in ['A3', 'B3', 'C3']:
    ws[cell].font = Font(name=FONT, size=11, bold=True)
    ws[cell].fill = SECTION_FILL

beneficiary_rows = [
    ('إجمالي ملفات المستفيدين النشطة 2025', '13000', 'الواقع كما ذكر CEO — تأكيد العدد الدقيق'),
    ('ملفات معتمدة', '', 'المعتمدة في النظام'),
    ('ملفات قيد المراجعة', '', 'لم تُعتمد بعد'),
    ('ملفات منتهية الصلاحية', '', 'تحتاج تجديد أو إغلاق'),
    ('', '', ''),
    ('SECTION', 'تفسير الفجوة مع ملف Excel (2,375)', ''),
    ('Excel "01 مايو 2026" يُظهر 2,375 ملف فقط', '', 'هل هذا فلتر؟ فئة معينة؟ أم تصدير جزئي؟'),
    ('شرح الفرق', '', '🟡 يُكتب هنا'),
    ('', '', ''),
    ('SECTION', 'توزيع فئات الملفات', ''),
    ('فئة أ', '', 'الأشد احتياجاً'),
    ('فئة ب', '', ''),
    ('فئة ج', '', ''),
    ('فئة د', '', 'الأقل احتياجاً نسبياً'),
    ('', '', ''),
    ('SECTION', 'المستهدفات السنوية للنمو', ''),
    ('الهدف 2026', '', 'كم ملف فريد بنهاية 2026؟'),
    ('الهدف 2027', '', ''),
    ('الهدف 2028', '', ''),
    ('الهدف 2029', '', ''),
    ('الهدف 2030', '', 'بدلاً من 13,000 الطموحي إن كان غير واقعي'),
]

row = 4
for a, b, c in beneficiary_rows:
    if a == '':
        row += 1
        continue
    if a == 'SECTION':
        ws.merge_cells(f'A{row}:C{row}')
        section(ws[f'A{row}'], b)
        ws.row_dimensions[row].height = 26
    else:
        label(ws[f'A{row}'], a)
        if b == '':
            input_cell(ws[f'B{row}'])
        else:
            correct(ws[f'B{row}'], b)
        note(ws[f'C{row}'], c)
        ws.row_dimensions[row].height = 24
    row += 1

# ════════════════════════════════════════════════════════
# الورقة 3: الكفالة
# ════════════════════════════════════════════════════════
ws = wb.create_sheet('2_الكفالة')
set_rtl(ws)
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 50

ws.merge_cells('A1:C1')
header(ws['A1'], 'برنامج كفالة الأيتام')

for cell, text in [('A3', 'البند'), ('B3', 'القيمة'), ('C3', 'ملاحظات')]:
    ws[cell].value = text
    ws[cell].font = Font(name=FONT, size=11, bold=True)
    ws[cell].fill = SECTION_FILL
    ws[cell].alignment = Alignment(horizontal='right', readingOrder=2)
    ws[cell].border = BORDER

kafala_rows = [
    ('SECTION', 'الواقع الفعلي 2025', ''),
    ('الأيتام المكفولون فعلياً (نشطون)', '800', 'كما ذكر CEO — تأكيد العدد'),
    ('الأيتام في قائمة الانتظار', '', 'مسجَّلون لكن بلا كفيل'),
    ('إجمالي الأيتام في الملفات (غير مكفولين)', '', 'موجودون في ملفات الأسر'),
    ('عدد الكفلاء النشطين 2025', '', 'كم شخص/جهة يكفلون؟'),
    ('متوسط مبلغ الكفالة الشهرية', '', 'بالريال'),
    ('', '', ''),
    ('SECTION', 'مشكلة "شبه الانتقاء"', ''),
    ('وصف المشكلة', '', '🟡 شرح طبيعة الانتقاء وأثره على الإنصاف'),
    ('عدد الحالات المتأثرة', '', 'تقدير'),
    ('الإجراء المقترح للمعالجة', '', '🟡 معايير اختيار جديدة؟ مراجعة شاملة؟'),
    ('', '', ''),
    ('SECTION', 'المستهدفات السنوية', ''),
    ('الهدف 2026', '', 'كم يتيم مكفول بنهاية 2026؟'),
    ('الهدف 2027', '', ''),
    ('الهدف 2028', '', ''),
    ('الهدف 2029', '', ''),
    ('الهدف 2030', '', 'بدلاً من 870 الأصلي'),
]

row = 4
for a, b, c in kafala_rows:
    if a == '':
        row += 1
        continue
    if a == 'SECTION':
        ws.merge_cells(f'A{row}:C{row}')
        section(ws[f'A{row}'], b)
        ws.row_dimensions[row].height = 26
    else:
        label(ws[f'A{row}'], a)
        if b == '':
            input_cell(ws[f'B{row}'])
        else:
            correct(ws[f'B{row}'], b)
        note(ws[f'C{row}'], c)
        ws.row_dimensions[row].height = 24
    row += 1

# ════════════════════════════════════════════════════════
# الورقة 4: الأرقام المالية
# ════════════════════════════════════════════════════════
ws = wb.create_sheet('3_الأرقام_المالية')
set_rtl(ws)
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 50

ws.merge_cells('A1:C1')
header(ws['A1'], 'الأرقام المالية الفعلية 2025')

for cell, text in [('A3', 'البند'), ('B3', 'القيمة (ريال)'), ('C3', 'ملاحظات')]:
    ws[cell].value = text
    ws[cell].font = Font(name=FONT, size=11, bold=True)
    ws[cell].fill = SECTION_FILL
    ws[cell].alignment = Alignment(horizontal='right', readingOrder=2)
    ws[cell].border = BORDER

financial_rows = [
    ('SECTION', 'الإيرادات 2025', ''),
    ('إجمالي الإيرادات 2025', '', '⚠️ المُدّعى 10.9M — CEO قال "مبالغ فيها جداً"'),
    ('التبرعات (أفراد)', '', ''),
    ('التبرعات (شركات/مؤسسات)', '', ''),
    ('الزكاة والصدقات', '', ''),
    ('الكفالات', '', ''),
    ('إيرادات مركز التدريب', '', '⚠️ كانت "50K تقديري"'),
    ('إيرادات استثمارية', '', ''),
    ('دعم حكومي', '', ''),
    ('أخرى', '', ''),
    ('', '', ''),
    ('SECTION', 'المصروفات 2025', ''),
    ('إجمالي المصروفات 2025', '', ''),
    ('برامج (مساعدات + تمكين)', '', ''),
    ('رواتب وأجور', '', ''),
    ('مصروفات تشغيلية', '', ''),
    ('مصروفات إدارية', '', ''),
    ('', '', ''),
    ('SECTION', 'المؤشرات المحسوبة', ''),
    ('الموارد الذاتية كقيمة (ريال)', '', ''),
    ('الموارد الذاتية كنسبة من الإيرادات', '', '⚠️ المُدّعى 13% — تأكيد؟'),
    ('عدد مصادر الإيرادات الفعّالة', '', '⚠️ المُدّعى 5 — تأكيد؟'),
    ('نسبة التعادل المالي', '', 'الإيرادات / المصروفات'),
    ('', '', ''),
    ('SECTION', 'المستهدفات السنوية الواقعية', ''),
    ('الإيرادات هدف 2026', '', ''),
    ('الإيرادات هدف 2027', '', ''),
    ('الإيرادات هدف 2028', '', ''),
    ('الإيرادات هدف 2029', '', ''),
    ('الإيرادات هدف 2030', '', '⚠️ المُدّعى 16.6M — CEO قال "مبالغ فيها"'),
    ('', '', ''),
    ('نسبة الموارد الذاتية 2026', '', ''),
    ('نسبة الموارد الذاتية 2027', '', ''),
    ('نسبة الموارد الذاتية 2028', '', ''),
    ('نسبة الموارد الذاتية 2029', '', ''),
    ('نسبة الموارد الذاتية 2030', '', '⚠️ المُدّعى 30% — تأكيد؟'),
]

row = 4
for a, b, c in financial_rows:
    if a == '':
        row += 1
        continue
    if a == 'SECTION':
        ws.merge_cells(f'A{row}:C{row}')
        section(ws[f'A{row}'], b)
        ws.row_dimensions[row].height = 26
    else:
        label(ws[f'A{row}'], a)
        if b == '':
            input_cell(ws[f'B{row}'])
        else:
            correct(ws[f'B{row}'], b)
        note(ws[f'C{row}'], c)
        ws.row_dimensions[row].height = 24
    row += 1

# ════════════════════════════════════════════════════════
# الورقة 5: الشراكات
# ════════════════════════════════════════════════════════
ws = wb.create_sheet('4_الشراكات')
set_rtl(ws)
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 30

ws.merge_cells('A1:F1')
header(ws['A1'], 'قائمة الشراكات الفعلية 2025 (الـ 18)')

headers = ['#', 'اسم الشريك', 'النوع', 'مجال الشراكة', 'سنة البدء', 'الهدف الذي تخدمه']
for i, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=i)
    c.value = h
    c.font = Font(name=FONT, size=11, bold=True)
    c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal='center', readingOrder=2)
    c.border = BORDER

# 18 صف لتعبئة الشراكات
for i in range(1, 19):
    r = 3 + i
    ws.cell(row=r, column=1).value = i
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=1).font = Font(name=FONT, size=11, bold=True)
    for col in range(2, 7):
        input_cell(ws.cell(row=r, column=col))

# نوع الشراكة - hint
ws['G3'] = 'أنواع الشراكة المقبولة:'
ws['G3'].font = Font(name=FONT, size=10, bold=True)
ws['G4'] = 'تمويلية · تمكينية · تشغيلية · تقنية · حكومية'
ws['G4'].font = Font(name=FONT, size=10, italic=True, color='595959')

# المستهدفات
row = 25
ws.merge_cells(f'A{row}:F{row}')
section(ws[f'A{row}'], 'المستهدفات السنوية للشراكات')
row += 1
for year in [2026, 2027, 2028, 2029, 2030]:
    ws.cell(row=row, column=2).value = f'إجمالي الشراكات الفعّالة بنهاية {year}'
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='right', readingOrder=2)
    ws.cell(row=row, column=2).border = BORDER
    ws.cell(row=row, column=2).font = Font(name=FONT, size=11)
    input_cell(ws.cell(row=row, column=3))
    row += 1

ws.cell(row=row, column=2).value = 'الهدف 2030 (إعادة نظر؟)'
ws.cell(row=row, column=2).font = Font(name=FONT, size=10, italic=True)
ws.cell(row=row, column=3).value = '60 (الأصلي)'
ws.cell(row=row, column=3).font = Font(name=FONT, size=10, italic=True, color='C00000')

# ════════════════════════════════════════════════════════
# الورقة 6: الكوادر
# ════════════════════════════════════════════════════════
ws = wb.create_sheet('5_الكوادر')
set_rtl(ws)
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 50

ws.merge_cells('A1:C1')
header(ws['A1'], 'الموظفون والمتطوعون')

for cell, text in [('A3', 'البند'), ('B3', 'القيمة'), ('C3', 'ملاحظات')]:
    ws[cell].value = text
    ws[cell].font = Font(name=FONT, size=11, bold=True)
    ws[cell].fill = SECTION_FILL
    ws[cell].alignment = Alignment(horizontal='right', readingOrder=2)
    ws[cell].border = BORDER

hr_rows = [
    ('SECTION', 'الموظفون 2025', ''),
    ('عدد الموظفين بدوام كامل', '', ''),
    ('عدد المتعاونين/جزئي', '', ''),
    ('متوسط ساعات التدريب/موظف 2025', '', 'الواقع الفعلي'),
    ('عدد الموظفين الذين أكملوا 20 ساعة تدريب', '', ''),
    ('', '', ''),
    ('SECTION', 'المتطوعون 2025', ''),
    ('إجمالي المتطوعين المُسجَّلين', '', ''),
    ('المتطوعون النشطون فعلياً', '', 'تعريف "نشط" = 24 ساعة سنوياً'),
    ('إجمالي ساعات التطوع 2025', '', ''),
    ('متوسط ساعات/متطوع نشط', '', 'محسوبة'),
    ('', '', ''),
    ('SECTION', 'مستهدفات الكوادر', ''),
    ('عدد الموظفين المدرّبين هدف 2030', '', 'بنسبة 90% — كم رقمياً؟'),
    ('عدد المتطوعين النشطين هدف 2030', '', '500 الأصلي — تأكيد؟'),
    ('ساعات التطوع هدف 2030', '', '30,000 الأصلي — تأكيد؟'),
]

row = 4
for a, b, c in hr_rows:
    if a == '':
        row += 1
        continue
    if a == 'SECTION':
        ws.merge_cells(f'A{row}:C{row}')
        section(ws[f'A{row}'], b)
        ws.row_dimensions[row].height = 26
    else:
        label(ws[f'A{row}'], a)
        if b == '':
            input_cell(ws[f'B{row}'])
        else:
            correct(ws[f'B{row}'], b)
        note(ws[f'C{row}'], c)
        ws.row_dimensions[row].height = 24
    row += 1

# ════════════════════════════════════════════════════════
# الورقة 7: المالكون والمبادرات
# ════════════════════════════════════════════════════════
ws = wb.create_sheet('6_المالكون_والمبادرات')
set_rtl(ws)
ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 25

ws.merge_cells('A1:D1')
header(ws['A1'], 'تصحيحات أصحاب المبادرات والأهداف')

# جزء أ: المبادرات
ws.merge_cells('A3:D3')
section(ws['A3'], 'أ. تصحيح أصحاب المبادرات')

for cell, text in [('A4', 'الكود'), ('B4', 'المبادرة'), ('C4', 'المسؤول الحالي (خطأ)'), ('D4', 'المسؤول الصحيح')]:
    ws[cell].value = text
    ws[cell].font = Font(name=FONT, size=11, bold=True)
    ws[cell].fill = PatternFill('solid', start_color='F4B084')
    ws[cell].alignment = Alignment(horizontal='right', readingOrder=2)
    ws[cell].border = BORDER

ini_rows = [
    ('INI-2026-005', 'كفالة الأيتام (200 يتيم)', 'طلال الحربي ❌', ''),  # input needed
    ('INI-2026-006', 'توزيع 1500 سلة شهرياً', 'خاتمة محرق ❌', 'طلال الحربي ✅'),
    ('INI-2026-018', 'تدريب الموظفين على AI', 'عبدالرحمن عقيل ❌', 'خليل هادي ✅'),
]

row = 5
for code, name, wrong, right in ini_rows:
    label(ws[f'A{row}'], code)
    label(ws[f'B{row}'], name)
    ws[f'C{row}'].value = wrong
    ws[f'C{row}'].font = Font(name=FONT, size=11, color='C00000')
    ws[f'C{row}'].fill = PatternFill('solid', start_color='FFE699')
    ws[f'C{row}'].alignment = Alignment(horizontal='right', readingOrder=2)
    ws[f'C{row}'].border = BORDER
    if right:
        correct(ws[f'D{row}'], right)
    else:
        input_cell(ws[f'D{row}'], 'من؟')
    ws.row_dimensions[row].height = 26
    row += 1

# مبادرات أخرى تحتاج مراجعة
row += 1
ws.merge_cells(f'A{row}:D{row}')
note(ws[f'A{row}'], 'مبادرات أخرى — هل تحتاج تصحيحاً؟ أضف رمز/اسم/مالك جديد')
row += 1
for _ in range(5):
    for col in range(1, 5):
        input_cell(ws.cell(row=row, column=col))
    row += 1

# جزء ب: الأهداف
row += 2
ws.merge_cells(f'A{row}:D{row}')
section(ws[f'A{row}'], 'ب. مراجعة أصحاب الأهداف الـ 8')
row += 1

for cell, text in [(f'A{row}', 'الكود'), (f'B{row}', 'الهدف'), (f'C{row}', 'المالك الحالي'), (f'D{row}', 'تأكيد/تصحيح')]:
    ws[cell].value = text
    ws[cell].font = Font(name=FONT, size=11, bold=True)
    ws[cell].fill = PatternFill('solid', start_color='F4B084')
    ws[cell].alignment = Alignment(horizontal='right', readingOrder=2)
    ws[cell].border = BORDER
row += 1

goal_rows = [
    ('STR-2026-003', 'تعميق الأثر الاجتماعي', 'خاتمة محرق'),
    ('STR-2026-004', 'تمكين المستفيدين', 'عبدالرحمن عقيل (مؤقت)'),
    ('STR-2026-006', 'تحسين تجربة المستفيد', 'خاتمة محرق'),
    ('STR-2026-007', 'الاستدامة المالية', 'نادية قلم'),
    ('STR-2026-012', 'التميز والحوكمة والجودة', 'ايلاف حسن'),
    ('STR-2026-013', 'التحول الرقمي', 'عبدالرحمن عقيل (مؤقت)'),
    ('STR-2026-016', 'منظومة الشراكات', 'فاطمة عقيبي'),
    ('STR-2026-017', 'تنمية الكوادر', 'خليل هادي'),
]

for code, name, current in goal_rows:
    label(ws[f'A{row}'], code)
    label(ws[f'B{row}'], name)
    label(ws[f'C{row}'], current)
    input_cell(ws[f'D{row}'], 'تأكيد أو اسم بديل')
    ws.row_dimensions[row].height = 24
    row += 1

# ════════════════════════════════════════════════════════
# حفظ
# ════════════════════════════════════════════════════════
output_path = r'C:\Users\abdu8\Documents\dev\qms\docs\plan-data-template-2026-04-30.xlsx'
wb.save(output_path)
print(f'✅ {output_path}')
