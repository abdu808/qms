import json
import os
import urllib.request
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
TODAY = "2026-05-02"
PLAN_DIR = ROOT / "ISO9001" / "الخطط والمشرات"
OUT_DIR = ROOT / "docs"
SOURCE_OPERATIONAL = PLAN_DIR / "الخطة_التشغيلية_2026_بـ4_محاور.xlsx"
OUTPUT_OPERATIONAL = PLAN_DIR / f"الخطة_التشغيلية_2026_بـ4_محاور_مع_ربط_المؤشرات_{TODAY}.xlsx"
OUTPUT_PACK = OUT_DIR / f"strategic-plan-correction-pack-{TODAY}.xlsx"
OUTPUT_MD = OUT_DIR / f"strategic-plan-correction-pack-{TODAY}.md"
OUTPUT_JSON = OUT_DIR / f"strategic-plan-correction-pack-{TODAY}.json"


def api_request(base_url, path, token=None, payload=None, method=None):
    headers = {"content-type": "application/json", "user-agent": "node"}
    if token:
        headers["authorization"] = f"Bearer {token}"
    body = None
    if payload is not None:
        body = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        base_url.rstrip("/") + path,
        data=body,
        headers=headers,
        method=method or ("POST" if payload is not None else "GET"),
    )
    with urllib.request.urlopen(req, timeout=45) as resp:
        return json.loads(resp.read().decode("utf-8"))


def login(base_url, email, password):
    return api_request(base_url, "/api/auth/login", payload={"email": email, "password": password})["token"]


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autofit(ws, max_width=55):
    for col_idx, col in enumerate(ws.columns, start=1):
        width = 12
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 3))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_sheet(ws, headers, rows):
    ws.append(headers)
    style_header(ws)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    autofit(ws)


def objective_indicator_rows(objectives, indicators):
    by_objective_id = {}
    for ind in indicators:
        by_objective_id.setdefault(ind.get("objectiveId"), []).append(ind)

    rows = []
    for obj in sorted(objectives, key=lambda x: x.get("code") or ""):
        linked = by_objective_id.get(obj.get("id"), [])
        if linked:
            for ind in sorted(linked, key=lambda x: x.get("code") or ""):
                rows.append({
                    "كود الهدف التشغيلي": obj.get("code"),
                    "الهدف التشغيلي": obj.get("title"),
                    "KPI في الهدف": obj.get("kpi"),
                    "كود المؤشر": ind.get("code"),
                    "اسم المؤشر المعتمد": ind.get("nameAr"),
                    "الوحدة": ind.get("unit") or "",
                    "المالك": (ind.get("owner") or {}).get("name") or (obj.get("owner") or {}).get("name") or "",
                    "الحكم": "مرتبط في النظام",
                })
        else:
            rows.append({
                "كود الهدف التشغيلي": obj.get("code"),
                "الهدف التشغيلي": obj.get("title"),
                "KPI في الهدف": obj.get("kpi"),
                "كود المؤشر": "",
                "اسم المؤشر المعتمد": "",
                "الوحدة": obj.get("unit") or "",
                "المالك": (obj.get("owner") or {}).get("name") or "",
                "الحكم": "يحتاج ربط مؤشر",
            })
    return rows


def mapping_stats(objectives, mapping_rows):
    objective_codes = {obj.get("code") for obj in objectives if obj.get("code")}
    mapped_codes = {row["كود الهدف التشغيلي"] for row in mapping_rows if row.get("كود المؤشر")}
    return {
        "mappedObjectives": len(mapped_codes),
        "totalObjectives": len(objective_codes),
        "linkedIndicators": sum(1 for row in mapping_rows if row.get("كود المؤشر")),
        "unmappedObjectives": sorted(objective_codes - mapped_codes),
    }


def initiative_decisions(file_initiatives, live_initiatives):
    live_by_code = {i["code"]: i for i in live_initiatives}
    file_by_code = {i["الكود"]: i for i in file_initiatives}
    codes = sorted(set(live_by_code) | set(file_by_code))
    rows = []
    for code in codes:
        live = live_by_code.get(code)
        local = file_by_code.get(code)
        if live and local:
            decision = "إبقاء ومطابقة"
            action = "لا إجراء جوهري"
        elif live and not local:
            decision = "إضافة للملف المعتمد"
            action = "تحديث ملف المبادرات لأن النظام يحتويها"
        else:
            decision = "حذف/أرشفة من الملف"
            action = "لا تنشأ في النظام ما لم يصدر قرار جديد"
        if code == "INI-2026-026":
            decision = "حذف/أرشفة من الملف"
            action = "تتوافق مع ملاحظة الإدارة: لا حاجة لشراكة TVTC بوجود معهد تابع للجمعية"
        if code == "INI-2026-032":
            decision = "إضافة للملف المعتمد"
            action = "مبادرة مهمة لتجربة المستفيد وربطها بقياس الرضا وSLA"
        if code == "INI-2026-033":
            decision = "إضافة للملف المعتمد"
            action = "مبادرة مباشرة لشهادة ISO 9001 وينبغي بقاؤها"
        rows.append({
            "الكود": code,
            "الاسم في النظام": live.get("name") if live else "",
            "الاسم في الملف": local.get("اسم المبادرة") if local else "",
            "الهدف": (live.get("goal") or {}).get("code") if live else local.get("الهدف الاستراتيجي"),
            "المالك": (live.get("owner") or {}).get("name") if live else local.get("المالك"),
            "الحالة": live.get("status") if live else local.get("الحالة"),
            "القرار": decision,
            "الإجراء": action,
        })
    return rows


def read_sheet_rows(path, required_header):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = None
    headers = None
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        values = [cell.value for cell in ws[row_idx]]
        if required_header in values:
            header_row = row_idx
            headers = values
            break
    if not header_row:
        raise ValueError(f"Header not found: {required_header}")
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        item = {str(h): row[idx] for idx, h in enumerate(headers) if h is not None}
        if item.get(required_header):
            rows.append(item)
    return rows


def read_kpi_catalog_rows():
    path = PLAN_DIR / "KPI_Catalog_v11_2026_2030.xlsx"
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["كتالوج المؤشرات"]
    headers = [cell.value for cell in ws[4]]
    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        item = {str(h): row[idx] for idx, h in enumerate(headers) if h is not None}
        if item.get("الكود"):
            rows.append(item)
    return rows


def enrich_operational_workbook(mapping_rows):
    wb = load_workbook(SOURCE_OPERATIONAL)
    ws = wb["الأهداف التشغيلية"]
    mapping = {r["كود الهدف التشغيلي"]: r for r in mapping_rows}
    header_row = 3
    existing_headers = [cell.value for cell in ws[header_row]]
    if "كود المؤشر المعتمد" not in existing_headers:
        insert_col = ws.max_column + 1
        ws.cell(header_row, insert_col, "كود المؤشر المعتمد")
        ws.cell(header_row, insert_col + 1, "اسم المؤشر المعتمد")
        ws.cell(header_row, insert_col + 2, "حالة الربط")
    else:
        insert_col = existing_headers.index("كود المؤشر المعتمد") + 1
    headers = [cell.value for cell in ws[header_row]]
    code_col = headers.index("الكود") + 1
    ind_code_col = headers.index("كود المؤشر المعتمد") + 1
    ind_name_col = headers.index("اسم المؤشر المعتمد") + 1
    status_col = headers.index("حالة الربط") + 1
    for row_idx in range(header_row + 1, ws.max_row + 1):
        obj_code = ws.cell(row_idx, code_col).value
        if not obj_code:
            continue
        mapped = mapping.get(str(obj_code))
        if mapped:
            ws.cell(row_idx, ind_code_col, mapped["كود المؤشر"])
            ws.cell(row_idx, ind_name_col, mapped["اسم المؤشر المعتمد"])
            ws.cell(row_idx, status_col, mapped["الحكم"])
    style_header(ws, header_row)
    autofit(ws)
    wb.save(OUTPUT_OPERATIONAL)


def main():
    base_url = os.environ.get("QMS_BASE_URL", "https://quality.aqiltech.sa")
    email = os.environ.get("QMS_EMAIL")
    password = os.environ.get("QMS_PASSWORD")
    if not email or not password:
        raise RuntimeError("QMS_EMAIL and QMS_PASSWORD are required.")

    token = login(base_url, email, password)
    objectives = api_request(base_url, "/api/objectives?limit=100", token)["items"]
    indicators = api_request(base_url, "/api/indicators?limit=100", token)["items"]
    live_initiatives = api_request(base_url, "/api/initiatives?limit=100", token)["items"]
    file_initiatives = read_sheet_rows(PLAN_DIR / "المبادرات_الـ21_2026_2030.xlsx", "الكود")
    file_kpis = read_kpi_catalog_rows()

    mapping_rows = objective_indicator_rows(objectives, indicators)
    stats = mapping_stats(objectives, mapping_rows)
    initiative_rows = initiative_decisions(file_initiatives, live_initiatives)
    baseline_rows = [
        {
            "كود المؤشر": ind.get("code"),
            "اسم المؤشر": ind.get("nameAr"),
            "خط الأساس الحالي": ind.get("baseline") if ind.get("baseline") is not None else "",
            "المالك": (ind.get("owner") or {}).get("name") or "",
            "الإجراء": "إبقاءه فارغًا مؤقتًا مع توثيق أول قياس" if ind.get("baseline") is None else "مكتمل",
        }
        for ind in sorted(indicators, key=lambda x: x.get("code") or "")
        if ind.get("baseline") is None
    ]
    file_baseline_rows = [
        {
            "كود المؤشر": row.get("الكود"),
            "اسم المؤشر": row.get("اسم المؤشر"),
            "خط الأساس في الملف": row.get("خط الأساس") or "",
            "المالك": row.get("المالك") or "",
            "الإجراء": "تعبئة خط الأساس في نسخة الملف أو توثيق موعد القياس الأول",
        }
        for row in file_kpis
        if row.get("خط الأساس") in (None, "", "—", "-")
    ]
    sensitive_rows = [
        {
            "البند": "OBJ-2026-022",
            "الملاحظة": "الوصول الخدمي لـ 13,000 أسرة",
            "القرار المقترح": "اعتماد التعريف: 13,000 خدمة/حالة موثقة تراكمية لا 13,000 أسرة فريدة",
            "الصياغة المقترحة": "الوصول الخدمي إلى 13,000 خدمة موثقة للأسر المستفيدة خلال 2026، مع تمييز عدد الأسر الفريدة في لوحة مستقلة",
        },
        {
            "البند": "OBJ-2026-029",
            "الملاحظة": "هدف توثيق 100% من السياسات والإجراءات لا يملك مؤشراً مستقلاً في النظام الحي",
            "القرار المقترح": "إنشاء مؤشر مستقل أو ربطه بمؤشر قائم بصفر وزن حتى لا يتكرر وزن الجودة",
            "الصياغة المقترحة": "نسبة السياسات والإجراءات المؤسسية المنشورة والمعتمدة في وحدة الوثائق من إجمالي القائمة المعتمدة",
        },
        {
            "البند": "INI-2026-005",
            "الملاحظة": "كفالة 870 يتيم",
            "القرار المقترح": "إعادة تعريف الرقم كمستفيدين متأثرين/مكفولين أو تخفيضه إلى نطاق الكفالة الفعلية",
            "الصياغة المقترحة": "تعزيز رعاية الأيتام عبر كفالات مباشرة وبرامج أثر نوعي، مع فصل عدد المكفولين فعلياً عن إجمالي الأيتام المسجلين",
        },
    ]

    enrich_operational_workbook(mapping_rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "ملخص"
    summary_rows = [
        {"البند": "تاريخ الإعداد", "القيمة": datetime.now().strftime("%Y-%m-%d %H:%M")},
        {"البند": "الأهداف التشغيلية المرتبطة بمؤشر واحد على الأقل", "القيمة": f"{stats['mappedObjectives']}/{stats['totalObjectives']}"},
        {"البند": "عدد المؤشرات المرتبطة بالأهداف", "القيمة": stats["linkedIndicators"]},
        {"البند": "مؤشرات بلا خط أساس رقمي", "القيمة": len(baseline_rows)},
        {"البند": "مؤشرات بلا خط أساس في ملف الكتالوج", "القيمة": len(file_baseline_rows)},
        {"البند": "أهداف تشغيلية بلا مؤشر مستقل في النظام", "القيمة": ", ".join(stats["unmappedObjectives"]) or "لا يوجد"},
        {"البند": "قرارات المبادرات", "القيمة": len(initiative_rows)},
        {"البند": "نسخة الخطة التشغيلية المحدثة", "القيمة": str(OUTPUT_OPERATIONAL.relative_to(ROOT)).replace('\\', '/')},
    ]
    write_sheet(ws, ["البند", "القيمة"], summary_rows)

    ws = wb.create_sheet("OBJ-IND")
    write_sheet(ws, ["كود الهدف التشغيلي", "الهدف التشغيلي", "KPI في الهدف", "كود المؤشر", "اسم المؤشر المعتمد", "الوحدة", "المالك", "الحكم"], mapping_rows)

    ws = wb.create_sheet("قرارات المبادرات")
    write_sheet(ws, ["الكود", "الاسم في النظام", "الاسم في الملف", "الهدف", "المالك", "الحالة", "القرار", "الإجراء"], initiative_rows)

    ws = wb.create_sheet("خطوط الأساس")
    write_sheet(ws, ["كود المؤشر", "اسم المؤشر", "خط الأساس الحالي", "المالك", "الإجراء"], baseline_rows)

    ws = wb.create_sheet("فجوات ملف الكتالوج")
    write_sheet(ws, ["كود المؤشر", "اسم المؤشر", "خط الأساس في الملف", "المالك", "الإجراء"], file_baseline_rows)

    ws = wb.create_sheet("حسم الصياغة")
    write_sheet(ws, ["البند", "الملاحظة", "القرار المقترح", "الصياغة المقترحة"], sensitive_rows)

    wb.save(OUTPUT_PACK)

    payload = {
        "objectiveIndicatorMap": mapping_rows,
        "initiativeDecisions": initiative_rows,
        "baselineGaps": baseline_rows,
        "fileBaselineGaps": file_baseline_rows,
        "sensitiveWording": sensitive_rows,
        "stats": stats,
        "outputs": {
            "pack": str(OUTPUT_PACK.relative_to(ROOT)).replace("\\", "/"),
            "operationalPlan": str(OUTPUT_OPERATIONAL.relative_to(ROOT)).replace("\\", "/"),
            "report": str(OUTPUT_MD.relative_to(ROOT)).replace("\\", "/"),
        },
    }
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    md = f"""# حزمة تصحيح الخطة الاستراتيجية والمؤشرات

**التاريخ:** {datetime.now().strftime("%Y-%m-%d %H:%M")}  
**مصدر المطابقة:** {base_url}

## النتيجة

تم تثبيت خريطة الربط بين الأهداف التشغيلية والمؤشرات المعتمدة. النظام الحي يحتوي الربط الرسمي عبر `Indicator.objectiveId`، وتم إنشاء نسخة محلية من الخطة التشغيلية تضيف أعمدة الربط حتى تصبح الوثيقة قابلة للمراجعة والاعتماد.

## الملفات الناتجة

- ملف الحزمة: `{str(OUTPUT_PACK.relative_to(ROOT)).replace("\\", "/")}`
- الخطة التشغيلية المحدثة بربط المؤشرات: `{str(OUTPUT_OPERATIONAL.relative_to(ROOT)).replace("\\", "/")}`
- ملف البيانات التفصيلي: `{str(OUTPUT_JSON.relative_to(ROOT)).replace("\\", "/")}`

## قرارات مطلوبة قبل أي تحديث إنتاجي جديد

1. اعتماد تعريف هدف `OBJ-2026-022` على أنه خدمات/حالات موثقة تراكمية، لا أسر فريدة، أو تعديل المستهدف.
2. إعادة صياغة `INI-2026-005` بحيث تفصل بين إجمالي الأيتام المسجلين وعدد المكفولين فعلياً.
3. حسم `OBJ-2026-029`: إنشاء مؤشر مستقل له أو إبقاؤه كمخرج تشغيلي تابع لمؤشر شهادة ISO.
4. حذف `INI-2026-026` من الملف المعتمد لأن النظام لم ينشئها ولأن ملاحظة الإدارة تشير إلى عدم الحاجة لها.
5. إضافة `INI-2026-032` و`INI-2026-033` إلى ملف المبادرات المعتمد لأنها موجودة في النظام وتخدم تجربة المستفيد وشهادة ISO.
6. تحديث خطوط الأساس الناقصة في ملف كتالوج المؤشرات المحلي؛ النظام الحي لا يظهر فجوة خط أساس حالياً.

## الحكم

لا أرى حاجة لتعديل فوري شامل في الإنتاج. الربط موجود لمعظم الأهداف، لكن `OBJ-2026-029` يحتاج قراراً: إما إنشاء مؤشر مستقل له، أو اعتباره مغطى ضمن مؤشر شهادة ISO/التوثيق العام دون وزن إضافي. التعديل المطلوب الآن وثائقي واعتمادي: جعل الملفات الرسمية تعكس ما هو موجود في النظام، ثم حسم الصياغات الحساسة قبل تجميد الخطة.
"""
    OUTPUT_MD.write_text(md, encoding="utf-8")

    print(json.dumps({
        "ok": True,
        "pack": str(OUTPUT_PACK.relative_to(ROOT)).replace("\\", "/"),
        "operationalPlan": str(OUTPUT_OPERATIONAL.relative_to(ROOT)).replace("\\", "/"),
        "report": str(OUTPUT_MD.relative_to(ROOT)).replace("\\", "/"),
        "mappedObjectives": stats["mappedObjectives"],
        "totalObjectives": stats["totalObjectives"],
        "linkedIndicators": stats["linkedIndicators"],
        "unmappedObjectives": stats["unmappedObjectives"],
        "baselineGaps": len(baseline_rows),
        "fileBaselineGaps": len(file_baseline_rows),
        "initiativeDecisions": len(initiative_rows),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
